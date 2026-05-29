# ==========================================
# C XML BR Engine
# COMPLETO + REV + HISTÓRICO + BANCO + REGISTROS
# ==========================================

import streamlit as st
import pdfplumber
import fitz
import re
import pandas as pd
from pathlib import Path
import zipfile
import tempfile
from ftfy import fix_text
import sqlite3
from datetime import datetime
from io import BytesIO
from zoneinfo import ZoneInfo
from openpyxl import load_workbook

# ==========================================
# CONFIG
# ==========================================

st.set_page_config(
    page_title="C XML BR Engine",
    layout="wide"
)

DB_PATH = "certificados.db"

# ==========================================
# CONEXÃO
# ==========================================

conn = sqlite3.connect(
    DB_PATH,
    check_same_thread=False
)

cursor = conn.cursor()

# ==========================================
# TABELAS
# ==========================================

cursor.execute("""
CREATE TABLE IF NOT EXISTS certificados (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    ip_bri TEXT UNIQUE,
    produto TEXT,
    ce_bri TEXT,
    familia TEXT,
    rev INTEGER,
    data_emissao TEXT,
    arquivo_pdf TEXT,
    data_cadastro TEXT,
    data_atualizacao TEXT
)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS itens (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    certificado_id INTEGER,
    ordem INTEGER,
    marca TEXT,
    modelo TEXT,
    nome TEXT,
    codigo TEXT
)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS historico_alteracoes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    ip_bri TEXT,
    rev_antiga INTEGER,
    rev_nova INTEGER,
    modelo TEXT,
    codigo TEXT,
    tipo_alteracao TEXT,
    campo_alterado TEXT,
    valor_antigo TEXT,
    valor_novo TEXT,
    arquivo_pdf TEXT,
    data_hora TEXT
)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS registros (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    cliente_base TEXT,
    fabrica TEXT,
    ce_bri TEXT,
    familia TEXT,
    registro TEXT,
    endereco_fabrica TEXT,
    arquivo_excel TEXT,
    data_cadastro TEXT
)
""")


cursor.execute("""
CREATE TABLE IF NOT EXISTS sistema5_clientes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    categoria TEXT,
    cliente_base TEXT,
    data_cadastro TEXT,
    UNIQUE(categoria, cliente_base)
)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS sistema5_fabricas (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    cliente_id INTEGER,
    fabrica TEXT,
    ce_bri TEXT,
    endereco_fabrica TEXT,
    data_cadastro TEXT,
    UNIQUE(cliente_id, fabrica)
)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS sistema5_arquivos (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    cliente_id INTEGER,
    fabrica_id INTEGER,
    tipo_processo TEXT,
    ip_processo TEXT,
    data_processo TEXT,
    arquivo_nome TEXT,
    data_upload TEXT
)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS sistema5_itens (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    arquivo_id INTEGER,
    cliente_base TEXT,
    categoria TEXT,
    fabrica TEXT,
    ce_bri TEXT,
    endereco_fabrica TEXT,
    tipo_processo TEXT,
    ip_processo TEXT,
    data_processo TEXT,
    familia TEXT,
    item TEXT,
    marca TEXT,
    modelo TEXT,
    nome TEXT,
    codigo TEXT,
    arquivo_nome TEXT,
    data_upload TEXT
)
""")


cursor.execute("""
CREATE TABLE IF NOT EXISTS ip_bri_familias (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    ce_bri TEXT,
    familia TEXT,
    ip_bri TEXT,
    observacao TEXT,
    data_cadastro TEXT,
    data_atualizacao TEXT,
    UNIQUE(ce_bri, familia)
)
""")

conn.commit()

# Índices para performance em buscas por campo mais usados
cursor.execute("CREATE INDEX IF NOT EXISTS idx_itens_codigo ON itens(codigo)")
cursor.execute("CREATE INDEX IF NOT EXISTS idx_itens_modelo ON itens(modelo)")
cursor.execute("CREATE INDEX IF NOT EXISTS idx_certificados_ce_bri ON certificados(ce_bri)")
cursor.execute("CREATE INDEX IF NOT EXISTS idx_certificados_ip_bri ON certificados(ip_bri)")
cursor.execute("CREATE INDEX IF NOT EXISTS idx_s5_itens_modelo ON sistema5_itens(modelo)")
cursor.execute("CREATE INDEX IF NOT EXISTS idx_s5_itens_arquivo_id ON sistema5_itens(arquivo_id)")
cursor.execute("CREATE INDEX IF NOT EXISTS idx_ip_bri_familias_ce_fam ON ip_bri_familias(ce_bri, familia)")
conn.commit()

# ==========================================
# FUNÇÕES GERAIS
# ==========================================


def clean(x):
    if x is None:
        return ""

    return re.sub(
        r"\s+",
        " ",
        str(x).replace("\n", " ")
    ).strip()




def normalizar_hifens_ref(valor):
    """
    Mantém o hífen como parte da referência, mas padroniza hífens unicode
    que visualmente parecem iguais.
    NÃO transforma SK-126 em SK126.
    """
    texto = clean(valor).upper()

    for h in ["‐", "‑", "‒", "–", "—", "−", "﹣", "－"]:
        texto = texto.replace(h, "-")

    texto = texto.replace("\u200b", "")
    texto = texto.replace("\u200c", "")
    texto = texto.replace("\u200d", "")
    texto = texto.replace("\ufeff", "")

    texto = re.sub(r"\s+", " ", texto).strip()

    return texto


def extrair_referencia_inicial_modelo(modelo):
    """
    Extrai a referência inicial do modelo.

    Ex:
    SK-126 - BRINQUEDO SANFONA -> SK-126
    SK-1261 - BRINQUEDO -> SK-1261
    SK126 - BRINQUEDO -> SK126
    """

    texto = normalizar_hifens_ref(modelo)

    partes = re.split(r"\s+-\s+", texto, maxsplit=1)

    if partes:
        return partes[0].strip()

    return texto.strip()


def modelo_corresponde_referencia_exata(modelo_banco, ref_busca):
    """
    Regra rígida da confirmação:

    SK-126 encontra:
    - SK-126
    - SK-126 - descrição

    SK-126 NÃO encontra:
    - SK126
    - SK-1261
    - SK-126A
    - SK-126 B
    """

    ref_buscada = normalizar_hifens_ref(ref_busca)
    ref_modelo = extrair_referencia_inicial_modelo(modelo_banco)

    if not ref_buscada or not ref_modelo:
        return False

    return ref_modelo == ref_buscada


def filtrar_resultado_por_referencia_exata(df, ref_busca):
    if df is None or df.empty:
        return pd.DataFrame()

    col_modelo = None

    if "MODELO" in df.columns:
        col_modelo = "MODELO"
    elif "modelo" in df.columns:
        col_modelo = "modelo"

    if not col_modelo:
        return pd.DataFrame()

    return df[
        df[col_modelo].apply(
            lambda modelo: modelo_corresponde_referencia_exata(modelo, ref_busca)
        )
    ].copy()


def corrigir_texto(texto):
    if texto is None:
        return ""

    return fix_text(str(texto))



def extrair_codigo_unico(codigo_raw):
    if not codigo_raw:
        return None

    codigo = re.split(r"[\/\n]", codigo_raw)[0]
    codigo = re.sub(r"\D", "", codigo)

    if not codigo or not codigo.isdigit():
        return None

    if len(codigo) < 8 or len(codigo) > 14:
        return None

    return codigo



def limitar_nome(nome):
    nome = clean(nome)

    if len(nome) > 200:
        nome = nome.replace("PRODUZIDO", "PROD.")
        nome = nome.replace("INDICATIVO", "IND.")
        nome = nome.replace("RESTRITIVO", "REST.")
        nome = nome[:200]

    return nome



def escape_xml(texto):
    return str(texto)\
        .replace("&", "&amp;")\
        .replace("<", "&lt;")\
        .replace(">", "&gt;")



def verificar_codigos_duplicados(df):
    duplicados = df[
        df.duplicated(
            subset=["CODIGO"],
            keep=False
        )
    ].copy()

    if not duplicados.empty:
        duplicados = duplicados.sort_values(by="CODIGO")

    return duplicados


# ==========================================
# REV
# ==========================================


def extrair_rev(texto):
    texto = corrigir_texto(texto)

    padroes = [
        r"\bREV\.?\s*:?\s*(\d+)",
        r"\bREVISÃO\s*:?\s*(\d+)",
        r"\bREVISAO\s*:?\s*(\d+)"
    ]

    for padrao in padroes:
        match = re.search(
            padrao,
            texto,
            flags=re.IGNORECASE
        )

        if match:
            return int(match.group(1))

    linhas = [clean(l) for l in texto.split("\n") if clean(l)]

    for i, linha in enumerate(linhas):
        if linha.upper() in ["REV", "REV.", "REVISÃO", "REVISAO"] and i + 1 < len(linhas):
            numero = re.sub(r"\D", "", linhas[i + 1])
            if numero:
                return int(numero)

    return 0


# ==========================================
# EXTRAÇÃO CERTIFICADO
# ==========================================


def extrair_familia_ip_bri(ip_bri):
    if not ip_bri:
        return None

    match = re.search(r"-([0-9]+)$", ip_bri)

    if not match:
        return None

    return str(int(match.group(1)))



def extrair_dados_certificado(pdf_path):
    doc = fitz.open(str(pdf_path))

    texto = ""

    for page in doc:
        texto += page.get_text() + "\n"

    texto = corrigir_texto(texto)

    ip_bri = None
    produto = None
    ce_bri = None
    data_emissao = None

    ip_match = re.search(
        r"IP-BRI-\d+\/\d+-\d+",
        texto
    )

    if ip_match:
        ip_bri = ip_match.group(0)

    ce_match = re.search(
        r"CE-BRI-[A-Z0-9\-]+",
        texto,
        flags=re.IGNORECASE
    )

    if ce_match:
        ce_bri = ce_match.group(0).upper()

    produto_match = re.search(
        r"Produto:\s*(.+)",
        texto
    )

    if produto_match:
        produto = clean(produto_match.group(1))

    emissao_match = re.search(
        r"Data de Emissão:\s*(\d{2}\/\d{2}\/\d{4})",
        texto,
        flags=re.IGNORECASE
    )

    if emissao_match:
        data_emissao = emissao_match.group(1)

    rev = extrair_rev(texto)
    familia = extrair_familia_ip_bri(ip_bri)

    return {
        "ip_bri": ip_bri,
        "produto": produto,
        "ce_bri": ce_bri,
        "rev": rev,
        "familia": familia,
        "data_emissao": data_emissao
    }


# ==========================================
# PARSE PDF
# ==========================================


def parse_pdf(pdf_path):
    rows = []

    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                for row in table:
                    if not row or len(row) < 6:
                        continue

                    ordem = clean(row[1])

                    if not re.fullmatch(r"\d{3}", ordem):
                        continue

                    ordem = int(ordem)

                    marca = clean(corrigir_texto(row[2]))
                    modelo = clean(corrigir_texto(row[3]))
                    nome = clean(corrigir_texto(row[4]))

                    codigo_raw = clean(corrigir_texto(row[5]))
                    codigo = extrair_codigo_unico(codigo_raw)

                    if not codigo:
                        continue

                    rows.append([
                        ordem,
                        marca,
                        modelo,
                        nome,
                        codigo
                    ])

    rows.sort(key=lambda x: x[0])

    return rows


# ==========================================
# XML
# ==========================================


def gerar_xml(df, sufixo):
    linhas = [
        '<?xml version="1.0" encoding="ISO-8859-1"?>',
        '<ArrayOfItemSolicitacao>'
    ]

    for _, r in df.iterrows():
        modelo = r["MODELO"].rstrip(".,") + sufixo

        linhas.append(f"""
<ItemSolicitacao>
<Marca>{escape_xml(r['MARCA'])}</Marca>
<Modelo>{escape_xml(modelo)}</Modelo>
<Nome>{escape_xml(limitar_nome(r['NOME']))}</Nome>
<CodigosBarras>
<Codigo>{r['CODIGO']}</Codigo>
</CodigosBarras>
</ItemSolicitacao>
""")

    linhas.append("</ArrayOfItemSolicitacao>")

    return "\n".join(linhas)


# ==========================================
# HISTÓRICO
# ==========================================


def registrar_historico(
    ip_bri,
    rev_antiga,
    rev_nova,
    modelo,
    codigo,
    tipo,
    campo,
    antigo,
    novo,
    arquivo
):
    cursor.execute("""
    INSERT INTO historico_alteracoes (
        ip_bri,
        rev_antiga,
        rev_nova,
        modelo,
        codigo,
        tipo_alteracao,
        campo_alterado,
        valor_antigo,
        valor_novo,
        arquivo_pdf,
        data_hora
    )
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        ip_bri,
        rev_antiga,
        rev_nova,
        modelo,
        codigo,
        tipo,
        campo,
        antigo,
        novo,
        arquivo,
        datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    ))

    conn.commit()



def comparar_itens(
    ip_bri,
    rev_antiga,
    rev_nova,
    antigos,
    novos,
    arquivo
):
    antigos_dict = {
        r[2]: r for r in antigos
    }

    novos_dict = {
        r[2]: r for r in novos
    }

    for modelo in antigos_dict:
        if modelo not in novos_dict:
            antigo = antigos_dict[modelo]

            registrar_historico(
                ip_bri,
                rev_antiga,
                rev_nova,
                antigo[2],
                antigo[4],
                "ITEM_REMOVIDO",
                "",
                f"MARCA: {antigo[1]} | MODELO: {antigo[2]} | NOME: {antigo[3]} | CÓDIGO: {antigo[4]}",
                "",
                arquivo
            )

    for modelo in novos_dict:
        if modelo not in antigos_dict:
            novo = novos_dict[modelo]

            registrar_historico(
                ip_bri,
                rev_antiga,
                rev_nova,
                novo[2],
                novo[4],
                "ITEM_NOVO",
                "",
                "",
                f"MARCA: {novo[1]} | MODELO: {novo[2]} | NOME: {novo[3]} | CÓDIGO: {novo[4]}",
                arquivo
            )

    for modelo in novos_dict:
        if modelo in antigos_dict:
            antigo = antigos_dict[modelo]
            novo = novos_dict[modelo]

            campos = {
                "MARCA": (antigo[1], novo[1]),
                "NOME": (antigo[3], novo[3]),
                "CODIGO": (antigo[4], novo[4])
            }

            for campo in campos:
                antigo_valor, novo_valor = campos[campo]

                if str(antigo_valor) != str(novo_valor):
                    registrar_historico(
                        ip_bri,
                        rev_antiga,
                        rev_nova,
                        modelo,
                        novo[4],
                        "CAMPO_ALTERADO",
                        campo,
                        antigo_valor,
                        novo_valor,
                        arquivo
                    )


# ==========================================
# SALVAR / ATUALIZAR
# ==========================================


def garantir_estrutura_banco():
    """
    Fonte única de verdade para estrutura do banco.
    Chamada UMA VEZ na inicialização via atualizar_familias_certificados().
    NÃO chamar por transação individual — custo desnecessário.
    """
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS certificados (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ip_bri TEXT UNIQUE,
        produto TEXT,
        ce_bri TEXT,
        rev INTEGER,
        data_emissao TEXT,
        arquivo_pdf TEXT,
        data_cadastro TEXT,
        data_atualizacao TEXT
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS itens (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        certificado_id INTEGER,
        ordem INTEGER,
        marca TEXT,
        modelo TEXT,
        nome TEXT,
        codigo TEXT
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS registros (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cliente_base TEXT,
        fabrica TEXT,
        ce_bri TEXT,
        familia TEXT,
        registro TEXT,
        arquivo_excel TEXT,
        data_cadastro TEXT
    )
    """)

    conn.commit()

    # Ajustes para bancos antigos (colunas que podem não existir)
    ajustes = [
        ("itens", "certificado_id", "INTEGER"),
        ("itens", "ordem", "INTEGER"),
        ("itens", "marca", "TEXT"),
        ("itens", "modelo", "TEXT"),
        ("itens", "nome", "TEXT"),
        ("itens", "codigo", "TEXT"),
        ("certificados", "produto", "TEXT"),
        ("certificados", "ce_bri", "TEXT"),
        ("certificados", "familia", "TEXT"),
        ("certificados", "rev", "INTEGER"),
        ("certificados", "data_emissao", "TEXT"),
        ("certificados", "arquivo_pdf", "TEXT"),
        ("certificados", "data_cadastro", "TEXT"),
        ("certificados", "data_atualizacao", "TEXT"),
        ("registros", "cliente_base", "TEXT"),
        ("registros", "ce_bri", "TEXT"),
        ("registros", "familia", "TEXT"),
        ("registros", "registro", "TEXT"),
        ("registros", "endereco_fabrica", "TEXT"),
        ("registros", "fabrica", "TEXT"),
        ("registros", "arquivo_excel", "TEXT"),
        ("registros", "data_cadastro", "TEXT"),
        ("sistema5_itens", "familia", "TEXT"),
        ("sistema5_itens", "item", "TEXT"),
        ("ip_bri_familias", "ce_bri", "TEXT"),
        ("ip_bri_familias", "familia", "TEXT"),
        ("ip_bri_familias", "ip_bri", "TEXT"),
        ("ip_bri_familias", "observacao", "TEXT"),
        ("ip_bri_familias", "data_cadastro", "TEXT"),
        ("ip_bri_familias", "data_atualizacao", "TEXT"),
    ]

    for tabela, coluna, tipo in ajustes:
        try:
            cursor.execute(f"ALTER TABLE {tabela} ADD COLUMN {coluna} {tipo}")
            conn.commit()
        except sqlite3.OperationalError:
            pass

    try:
        cursor.execute("UPDATE registros SET cliente_base = 'BOLSA' WHERE cliente_base IS NULL OR cliente_base = ''")
        conn.commit()
    except Exception:
        pass


def reparar_tabela_itens_se_necessario():
    """
    Verifica e repara a estrutura da tabela itens se necessário.
    Chamada UMA VEZ na inicialização. Nunca chamar por item inserido.
    """
    obrigatorias = {
        "id",
        "certificado_id",
        "ordem",
        "marca",
        "modelo",
        "nome",
        "codigo"
    }

    try:
        info = cursor.execute("PRAGMA table_info(itens)").fetchall()
        existentes = {linha[1] for linha in info}
    except Exception:
        existentes = set()

    if obrigatorias.issubset(existentes):
        return

    # Cria uma tabela nova correta e tenta preservar o que for possível
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS itens_corrigida (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        certificado_id INTEGER,
        ordem INTEGER,
        marca TEXT,
        modelo TEXT,
        nome TEXT,
        codigo TEXT
    )
    """)

    colunas_comuns = [c for c in ["id", "certificado_id", "ordem", "marca", "modelo", "nome", "codigo"] if c in existentes]

    if colunas_comuns:
        cols = ", ".join(colunas_comuns)
        try:
            cursor.execute(f"INSERT OR IGNORE INTO itens_corrigida ({cols}) SELECT {cols} FROM itens")
        except Exception:
            pass

    try:
        cursor.execute("DROP TABLE itens")
    except Exception:
        pass

    cursor.execute("ALTER TABLE itens_corrigida RENAME TO itens")
    conn.commit()


def inserir_item_seguro(certificado_id, r):
    cursor.execute("""
    INSERT INTO itens (
        certificado_id,
        ordem,
        marca,
        modelo,
        nome,
        codigo
    )
    VALUES (?, ?, ?, ?, ?, ?)
    """, (
        certificado_id,
        r[0],
        r[1],
        r[2],
        r[3],
        r[4]
    ))


def salvar_ou_atualizar_certificado(
    dados_certificado,
    rows,
    nome_arquivo
):
    ip_bri = dados_certificado["ip_bri"]
    rev_nova = dados_certificado["rev"]

    if not ip_bri:
        return "erro", "IP-BRI não encontrado"

    existente = cursor.execute("""
    SELECT id, rev
    FROM certificados
    WHERE ip_bri = ?
    """, (ip_bri,)).fetchone()

    if not existente:
        cursor.execute("""
        INSERT INTO certificados (
            ip_bri,
            produto,
            ce_bri,
            familia,
            rev,
            data_emissao,
            arquivo_pdf,
            data_cadastro,
            data_atualizacao
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            ip_bri,
            dados_certificado["produto"],
            dados_certificado["ce_bri"],
            dados_certificado.get("familia"),
            rev_nova,
            dados_certificado["data_emissao"],
            nome_arquivo,
            datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        ))

        conn.commit()

        certificado_id = cursor.lastrowid

        for r in rows:
            cursor.execute("""
            INSERT INTO itens (
                certificado_id,
                ordem,
                marca,
                modelo,
                nome,
                codigo
            )
            VALUES (?, ?, ?, ?, ?, ?)
            """, (
                certificado_id,
                r[0],
                r[1],
                r[2],
                r[3],
                r[4]
            ))

        conn.commit()

        registrar_historico(
            ip_bri,
            None,
            rev_nova,
            "",
            "",
            "CERTIFICADO_NOVO",
            "",
            "",
            f"Novo certificado com {len(rows)} itens",
            nome_arquivo
        )

        return "novo", "Novo certificado salvo"

    certificado_id = existente[0]
    rev_antiga = int(existente[1] or 0)

    # Verifica se o certificado existe, mas ficou sem itens no banco.
    # Nesse caso, mesmo com REV igual ou menor, o sistema preenche os itens sem apagar dados úteis.
    qtd_itens_existentes = cursor.execute("""
    SELECT COUNT(*)
    FROM itens
    WHERE certificado_id = ?
    """, (certificado_id,)).fetchone()[0]

    if rev_nova <= rev_antiga and qtd_itens_existentes > 0:
        registrar_historico(
            ip_bri,
            rev_antiga,
            rev_nova,
            "",
            "",
            "REV_IGNORADA",
            "",
            f"REV banco: {rev_antiga}",
            f"REV enviada: {rev_nova}",
            nome_arquivo
        )

        return "ignorado", "REV menor ou igual. Banco não alterado."

    if rev_nova <= rev_antiga and qtd_itens_existentes == 0:
        cursor.execute("""
        UPDATE certificados
        SET
            produto = ?,
            ce_bri = ?,
            familia = ?,
            rev = ?,
            data_emissao = ?,
            arquivo_pdf = ?,
            data_atualizacao = ?
        WHERE id = ?
        """, (
            dados_certificado["produto"],
            dados_certificado["ce_bri"],
            dados_certificado.get("familia"),
            rev_nova,
            dados_certificado["data_emissao"],
            nome_arquivo,
            datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            certificado_id
        ))

        for r in rows:
            inserir_item_seguro(certificado_id, r)

        conn.commit()

        registrar_historico(
            ip_bri,
            rev_antiga,
            rev_nova,
            "",
            "",
            "ITENS_REINSERIDOS",
            "",
            "Certificado existia sem itens vinculados",
            f"{len(rows)} itens inseridos na REV {rev_nova}",
            nome_arquivo
        )

        return "corrigido", f"Certificado já existia, mas estava sem itens. {len(rows)} itens foram inseridos."

    antigos = cursor.execute("""
    SELECT
        ordem,
        marca,
        modelo,
        nome,
        codigo
    FROM itens
    WHERE certificado_id = ?
    """, (certificado_id,)).fetchall()

    comparar_itens(
        ip_bri,
        rev_antiga,
        rev_nova,
        antigos,
        rows,
        nome_arquivo
    )

    cursor.execute("""
    DELETE FROM itens
    WHERE certificado_id = ?
    """, (certificado_id,))

    cursor.execute("""
    UPDATE certificados
    SET
        produto = ?,
        ce_bri = ?,
        familia = ?,
        rev = ?,
        data_emissao = ?,
        arquivo_pdf = ?,
        data_atualizacao = ?
    WHERE id = ?
    """, (
        dados_certificado["produto"],
        dados_certificado["ce_bri"],
        dados_certificado.get("familia"),
        rev_nova,
        dados_certificado["data_emissao"],
        nome_arquivo,
        datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        certificado_id
    ))

    for r in rows:
        inserir_item_seguro(certificado_id, r)

    conn.commit()

    registrar_historico(
        ip_bri,
        rev_antiga,
        rev_nova,
        "",
        "",
        "CERTIFICADO_ATUALIZADO",
        "",
        f"REV {rev_antiga}",
        f"REV {rev_nova}",
        nome_arquivo
    )

    return "atualizado", f"Atualizado REV {rev_antiga} → {rev_nova}"


# ==========================================
# EXPORTAR
# ==========================================


def exportar_todos_itens():
    return pd.read_sql_query("""
    SELECT
        i.marca AS marca,
        i.modelo AS modelo,
        i.nome AS nome,
        i.codigo AS codigo,
        c.ip_bri AS ip_bri,
        c.ce_bri AS ce_bri,
        r.registro AS registro,
        c.familia AS fam,
        r.fabrica AS fabrica,
        c.rev AS rev,
        c.produto AS produto,
        c.data_emissao AS data_emissao,
        i.ordem AS ordem
    FROM itens i
    INNER JOIN certificados c
    ON c.id = i.certificado_id
    LEFT JOIN registros r
    ON UPPER(r.ce_bri) = UPPER(c.ce_bri)
    AND r.familia = c.familia
    ORDER BY i.marca, i.modelo, c.ip_bri, i.ordem
    """, conn)


# ==========================================
# REGISTROS EXCEL
# ==========================================


def normalizar_familia(valor):
    if pd.isna(valor):
        return ""

    texto = str(valor).strip()

    if texto.endswith(".0"):
        texto = texto[:-2]

    numeros = re.findall(r"[0-9]+", texto)

    if numeros:
        return str(int(numeros[0]))

    return ""



def normalizar_registro(valor):
    if pd.isna(valor):
        return ""

    texto = str(valor).strip()

    if texto.endswith(".0"):
        texto = texto[:-2]

    return texto



def parece_registro(valor):
    texto = normalizar_registro(valor)

    if not texto:
        return False

    return bool(re.search(r"[0-9]+[ ]*/[ ]*[0-9]+", texto))



def extrair_ce_bri_da_aba(aba):
    match = re.search(
        r"CE-BRI-[A-Z0-9\-]+",
        str(aba),
        flags=re.IGNORECASE
    )

    if match:
        return match.group(0).upper()

    return None



def ler_registros_aba(excel_file, aba):
    bruto = pd.read_excel(
        excel_file,
        sheet_name=aba,
        header=None,
        dtype=object
    )

    registros = []

    for i, row in bruto.iterrows():
        valores = [str(v).strip().upper() for v in row.values]

        cols_familia = [
            idx for idx, valor in enumerate(valores)
            if "FAMILIA" in valor or "FAMÍLIA" in valor
        ]

        cols_registro = [
            idx for idx, valor in enumerate(valores)
            if "REGISTRO" in valor
        ]

        if not cols_familia or not cols_registro:
            continue

        for col_familia in cols_familia:
            registros_direita = [c for c in cols_registro if c > col_familia]

            if registros_direita:
                col_registro = registros_direita[0]
            else:
                col_registro = min(cols_registro, key=lambda c: abs(c - col_familia))

            for j in range(i + 1, len(bruto)):
                linha_atual = [str(v).strip().upper() for v in bruto.iloc[j].values]
                linha_texto = " ".join([v for v in linha_atual if v and v.lower() != "nan"])

                if ("FAMILIA" in linha_texto or "FAMÍLIA" in linha_texto) and "REGISTRO" in linha_texto:
                    break

                familia_raw = bruto.iat[j, col_familia] if col_familia < bruto.shape[1] else None
                registro_raw = bruto.iat[j, col_registro] if col_registro < bruto.shape[1] else None

                familia = normalizar_familia(familia_raw)
                registro = normalizar_registro(registro_raw)

                if familia and registro and parece_registro(registro):
                    registros.append({
                        "FAMILIA": familia,
                        "REGISTRO": registro,
                        "FABRICA": aba,
                        "CE_BRI": extrair_ce_bri_da_aba(aba)
                    })

    if not registros:
        for i, row in bruto.iterrows():
            for col in range(bruto.shape[1] - 1):
                familia = normalizar_familia(bruto.iat[i, col])
                registro = normalizar_registro(bruto.iat[i, col + 1])

                if familia and registro and parece_registro(registro):
                    registros.append({
                        "FAMILIA": familia,
                        "REGISTRO": registro,
                        "FABRICA": aba,
                        "CE_BRI": extrair_ce_bri_da_aba(aba)
                    })

    if not registros:
        return None

    dados = pd.DataFrame(registros)

    dados = dados.drop_duplicates(
        subset=["FAMILIA", "REGISTRO", "FABRICA", "CE_BRI"]
    )

    dados["FAMILIA_NUM"] = pd.to_numeric(dados["FAMILIA"], errors="coerce")
    dados = dados.sort_values(
        by=["FAMILIA_NUM", "REGISTRO"]
    ).drop(columns=["FAMILIA_NUM"])

    return dados



def salvar_registros_no_banco(banco_registros, arquivo_nome, cliente_base, enderecos_fabricas=None):
    if banco_registros is None or banco_registros.empty:
        return 0

    cliente_base = clean(cliente_base).upper()
    enderecos_fabricas = enderecos_fabricas or {}

    # Apaga somente os registros da base atual.
    # Exemplo: salvar BOLSA não apaga MOHNISH.
    cursor.execute("DELETE FROM registros WHERE UPPER(cliente_base) = UPPER(?)", (cliente_base,))

    total = 0

    for _, r in banco_registros.iterrows():
        fabrica = r.get("FABRICA", "")
        endereco = enderecos_fabricas.get(str(fabrica), "")

        cursor.execute("""
        INSERT INTO registros (
            cliente_base,
            fabrica,
            ce_bri,
            familia,
            registro,
            endereco_fabrica,
            arquivo_excel,
            data_cadastro
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            cliente_base,
            fabrica,
            r.get("CE_BRI", ""),
            str(r.get("FAMILIA", "")),
            r.get("REGISTRO", ""),
            endereco,
            arquivo_nome,
            datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        ))
        total += 1

    conn.commit()
    return total


def exportar_registros_banco():
    return pd.read_sql_query("""
    SELECT
        cliente_base,
        fabrica,
        ce_bri,
        familia,
        registro,
        endereco_fabrica,
        arquivo_excel,
        data_cadastro
    FROM registros
    ORDER BY cliente_base, fabrica, CAST(familia AS INTEGER), registro
    """, conn)



def buscar_registro_por_certificado(ce_bri, familia):
    if not ce_bri or not familia:
        return pd.DataFrame()

    return pd.read_sql_query("""
    SELECT
        fabrica,
        ce_bri,
        familia,
        registro,
        endereco_fabrica,
        arquivo_excel,
        data_cadastro
    FROM registros
    WHERE UPPER(ce_bri) = UPPER(?)
    AND familia = ?
    ORDER BY fabrica, familia, registro
    """, conn, params=(ce_bri, str(int(familia))))


def gerar_backup_geral_zip():
    buffer = BytesIO()

    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
        data_hora = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

        if Path(DB_PATH).exists():
            zipf.write(DB_PATH, f"backup_db/certificados_{data_hora}.db")

        try:
            banco_completo = exportar_todos_itens()
            zipf.writestr(
                f"csv/banco_completo_{data_hora}.csv",
                banco_completo.to_csv(index=False, sep=";").encode("ISO-8859-1", errors="replace")
            )
        except Exception:
            pass

        try:
            registros = exportar_registros_banco()
            zipf.writestr(
                f"csv/registros_{data_hora}.csv",
                registros.to_csv(index=False, sep=";").encode("ISO-8859-1", errors="replace")
            )
        except Exception:
            pass

        try:
            certificados = pd.read_sql_query("SELECT * FROM certificados ORDER BY ip_bri", conn)
            zipf.writestr(
                f"csv/certificados_{data_hora}.csv",
                certificados.to_csv(index=False, sep=";").encode("ISO-8859-1", errors="replace")
            )
        except Exception:
            pass

        try:
            itens = pd.read_sql_query("SELECT * FROM itens ORDER BY certificado_id, ordem", conn)
            zipf.writestr(
                f"csv/itens_puros_{data_hora}.csv",
                itens.to_csv(index=False, sep=";").encode("ISO-8859-1", errors="replace")
            )
        except Exception:
            pass

        try:
            historico = pd.read_sql_query("SELECT * FROM historico_alteracoes ORDER BY id DESC", conn)
            zipf.writestr(
                f"csv/historico_{data_hora}.csv",
                historico.to_csv(index=False, sep=";").encode("ISO-8859-1", errors="replace")
            )
        except Exception:
            pass

        try:
            resumo = f"""BACKUP GERAL C XML BR ENGINE
Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}

Arquivos incluídos:
- certificados.db
- banco_completo.csv
- registros.csv
- certificados.csv
- itens_puros.csv
- historico.csv
"""
            zipf.writestr("LEIA-ME.txt", resumo.encode("ISO-8859-1", errors="replace"))
        except Exception:
            pass

    buffer.seek(0)
    return buffer



def aviso_backup_diario():
    try:
        hora_sp = datetime.now(ZoneInfo("America/Sao_Paulo")).hour
        if hora_sp >= 18:
            st.sidebar.error("⚠️ Faça o backup geral antes de sair. O Streamlit pode reiniciar e apagar os dados.")
    except Exception:
        pass



def atualizar_familias_certificados():
    """
    Executada UMA VEZ na inicialização.
    Centraliza: estrutura do banco, reparo da tabela itens e famílias faltantes.
    """
    try:
        garantir_estrutura_banco()
    except Exception:
        pass

    try:
        reparar_tabela_itens_se_necessario()
    except Exception:
        pass

    try:
        certificados = cursor.execute("SELECT id, ip_bri FROM certificados").fetchall()
        for cert_id, ip_bri_atual in certificados:
            fam = extrair_familia_ip_bri(ip_bri_atual)
            if fam:
                cursor.execute("UPDATE certificados SET familia = ? WHERE id = ?", (fam, cert_id))
        conn.commit()
    except Exception:
        pass


atualizar_familias_certificados()

# ==========================================
# PREENCHIMENTO DE CONFIRMAÇÃO
# ==========================================


def cor_rgb(cell):
    try:
        fill = cell.fill
        if fill is None or fill.fill_type is None:
            return None

        color = fill.fgColor

        if color is None:
            return None

        rgb = color.rgb

        if not rgb:
            return None

        rgb = str(rgb).replace("#", "")

        if len(rgb) == 8:
            rgb = rgb[2:]

        if len(rgb) != 6:
            return None

        return tuple(int(rgb[i:i+2], 16) for i in (0, 2, 4))
    except Exception:
        return None



def eh_amarelo(cell):
    rgb = cor_rgb(cell)
    if not rgb:
        return False

    r, g, b = rgb
    return r >= 180 and g >= 150 and b <= 140



def eh_verde(cell):
    rgb = cor_rgb(cell)
    if not rgb:
        return False

    r, g, b = rgb
    return g >= 120 and r <= 180 and b <= 180 and g >= r



def eh_azul(cell):
    rgb = cor_rgb(cell)
    if not rgb:
        return False

    r, g, b = rgb
    return b >= 120 and b > r and b >= g



def normalizar_cabecalho(valor):
    texto = clean(valor).upper()
    texto = texto.replace("CÓDIGO", "CODIGO")
    texto = texto.replace("COD.", "CODIGO")
    texto = texto.replace("CÓD.", "CODIGO")
    texto = texto.replace("IP DO PROCESSO", "IP_PROCESSO")
    texto = texto.replace("IP_PROCESSO", "IP_PROCESSO")
    texto = texto.replace("IP-PROCESSO", "IP_PROCESSO")
    texto = texto.replace("IP PROCESSO", "IP_PROCESSO")
    texto = texto.replace("IP-BRI", "IP_BRI")
    texto = texto.replace("CE-BRI", "CE_BRI")
    texto = texto.replace("ENDEREÇO", "ENDERECO")
    texto = texto.replace("END.", "ENDERECO")
    texto = texto.replace("ENDERECO_DA_FABRICA", "ENDERECO")
    texto = texto.replace("ENDEREÇO_DA_FABRICA", "ENDERECO")
    texto = texto.replace("ENDERECO_FABRICA", "ENDERECO")
    texto = texto.replace("ENDEREÇO_FABRICA", "ENDERECO")
    texto = texto.replace(" ", "_")
    texto = texto.replace("ENDERECO_DA_FABRICA", "ENDERECO")
    texto = texto.replace("ENDERECO_FABRICA", "ENDERECO")
    return texto



def buscar_item_confirmacao(valor_ref):
    ref = clean(valor_ref)

    if not ref:
        return None

    item_s5 = buscar_item_sistema5_confirmacao(ref)

    if item_s5:
        return item_s5

    # Busca candidatos no certificado oficial, mas só aceita se a referência inicial for exatamente igual.
    resultado = pd.read_sql_query("""
    SELECT
        i.marca AS MARCA,
        i.modelo AS MODELO,
        i.nome AS NOME,
        i.codigo AS CODIGO,
        c.ip_bri AS IP_BRI,
        c.ce_bri AS CE_BRI,
        r.registro AS REGISTRO,
        r.endereco_fabrica AS ENDERECO
    FROM itens i
    INNER JOIN certificados c
    ON c.id = i.certificado_id
    LEFT JOIN registros r
    ON UPPER(r.ce_bri) = UPPER(c.ce_bri)
    AND r.familia = c.familia
    WHERE UPPER(TRIM(i.modelo)) = UPPER(TRIM(?))
       OR UPPER(TRIM(i.modelo)) LIKE UPPER(TRIM(?) || ' -%')
       OR UPPER(TRIM(i.modelo)) LIKE UPPER(TRIM(?) || '%')
    ORDER BY c.rev DESC, c.ip_bri DESC
    LIMIT 100
    """, conn, params=(ref, ref, ref))

    resultado = filtrar_resultado_por_referencia_exata(resultado, ref)

    if resultado.empty:
        return None

    return resultado.iloc[0].to_dict()


def descobrir_cabecalho_coluna(ws, row_idx, col_idx):
    # Primeiro procura cabeçalho azul acima da célula amarela
    for r in range(row_idx - 1, 0, -1):
        cell = ws.cell(r, col_idx)
        valor = clean(cell.value)

        if valor and (eh_azul(cell) or normalizar_cabecalho(valor) in ["MARCA", "MODELO", "NOME", "CODIGO", "IP_BRI", "CE_BRI", "REGISTRO", "ENDERECO", "FAMILIA", "ITEM", "TIPO_PROCESSO", "DATA_PROCESSO", "ARQUIVO_ORIGEM", "IP_PROCESSO"]):
            return normalizar_cabecalho(valor)

    # Plano B: procura qualquer texto de cabeçalho acima
    for r in range(row_idx - 1, 0, -1):
        valor = clean(ws.cell(r, col_idx).value)
        cab = normalizar_cabecalho(valor)

        if cab in ["MARCA", "MODELO", "NOME", "CODIGO", "IP_BRI", "CE_BRI", "REGISTRO", "ENDERECO", "FAMILIA", "ITEM", "TIPO_PROCESSO", "DATA_PROCESSO", "ARQUIVO_ORIGEM", "IP_PROCESSO"]:
            return cab

    return ""



def preencher_excel_confirmacao(uploaded_file):
    wb = load_workbook(uploaded_file)

    preenchidos = 0
    nao_encontrados = []

    campos_validos = ["MARCA", "MODELO", "NOME", "CODIGO", "IP_BRI", "CE_BRI", "REGISTRO", "ENDERECO", "FAMILIA", "ITEM", "TIPO_PROCESSO", "DATA_PROCESSO", "ARQUIVO_ORIGEM", "IP_PROCESSO"]

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            refs = []

            # REGRA OFICIAL:
            # Qualquer célula verde da linha é considerada referência.
            # Não importa o nome da coluna.
            # Não precisa existir coluna chamada REF.
            for cell in row:
                if eh_verde(cell) and clean(cell.value):
                    refs.append(clean(cell.value))

            if not refs:
                continue

            item = None

            # Se houver mais de uma célula verde na linha, tenta na ordem em que aparece.
            for ref in refs:
                item = buscar_item_confirmacao(ref)
                if item:
                    break

            if not item:
                nao_encontrados.extend(refs)
                continue

            for cell in row:
                if not eh_amarelo(cell):
                    continue

                campo = descobrir_cabecalho_coluna(ws, cell.row, cell.column)

                if campo not in campos_validos:
                    continue

                valor = item.get(campo)

                if valor is None:
                    continue

                valor = str(valor).strip()

                if valor == "" or valor.lower() == "nan":
                    continue

                cell.value = valor
                preenchidos += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output, preenchidos, nao_encontrados


def atualizar_endereco_fabrica_existente(cliente_base, fabrica, novo_endereco):
    cliente_base = clean(cliente_base).upper()
    fabrica = clean(fabrica)
    novo_endereco = clean(novo_endereco)

    cursor.execute("""
    UPDATE registros
    SET endereco_fabrica = ?
    WHERE UPPER(cliente_base) = UPPER(?)
    AND fabrica = ?
    """, (
        novo_endereco,
        cliente_base,
        fabrica
    ))

    conn.commit()
    return cursor.rowcount






# ==========================================
# IP-BRI MANUAL POR FAMÍLIA
# ==========================================

def normalizar_ip_bri_manual(valor):
    texto = clean(valor).upper()

    if not texto:
        return ""

    if texto.startswith("IP-BRI-"):
        return texto

    if texto.startswith("IP-"):
        numero = texto.replace("IP-", "")
        return f"IP-BRI-{numero}"

    return f"IP-BRI-{texto}"


def salvar_ip_bri_familia(ce_bri, familia, ip_bri, observacao=""):
    ce_bri = clean(ce_bri).upper()
    familia = normalizar_familia(familia)
    ip_bri = normalizar_ip_bri_manual(ip_bri)
    observacao = clean(observacao)

    if not ce_bri or not familia or not ip_bri:
        return False, "Informe CE-BRI, família e IP-BRI."

    # Bloqueia o mesmo IP-BRI em outra família/CE-BRI.
    duplicado = cursor.execute("""
    SELECT ce_bri, familia
    FROM ip_bri_familias
    WHERE UPPER(ip_bri) = UPPER(?)
    AND NOT (
        UPPER(ce_bri) = UPPER(?)
        AND familia = ?
    )
    LIMIT 1
    """, (ip_bri, ce_bri, familia)).fetchone()

    if duplicado:
        return False, (
            f"Este IP-BRI já está cadastrado em outro vínculo: "
            f"{duplicado[0]} / família {duplicado[1]}. "
            f"Edite ou exclua o cadastro antigo antes de salvar."
        )

    existente = cursor.execute("""
    SELECT id
    FROM ip_bri_familias
    WHERE UPPER(ce_bri) = UPPER(?)
    AND familia = ?
    """, (ce_bri, familia)).fetchone()

    agora_txt = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    if existente:
        cursor.execute("""
        UPDATE ip_bri_familias
        SET ip_bri = ?, observacao = ?, data_atualizacao = ?
        WHERE id = ?
        """, (ip_bri, observacao, agora_txt, existente[0]))
    else:
        cursor.execute("""
        INSERT INTO ip_bri_familias (
            ce_bri, familia, ip_bri, observacao, data_cadastro, data_atualizacao
        )
        VALUES (?, ?, ?, ?, ?, ?)
        """, (ce_bri, familia, ip_bri, observacao, agora_txt, agora_txt))

    conn.commit()
    return True, f"IP-BRI {ip_bri} salvo para {ce_bri} / família {familia}."

def buscar_ip_bri_manual_familia(ce_bri, familia):
    ce_bri = clean(ce_bri).upper()
    familia = normalizar_familia(familia)

    if not ce_bri or not familia:
        return ""

    resultado = pd.read_sql_query("""
    SELECT ip_bri
    FROM ip_bri_familias
    WHERE UPPER(ce_bri) = UPPER(?)
    AND familia = ?
    ORDER BY id DESC
    LIMIT 1
    """, conn, params=(ce_bri, familia))

    if resultado.empty:
        return ""

    return clean(resultado.iloc[0]["ip_bri"])


def listar_ip_bri_pendentes_sistema5():
    return pd.read_sql_query("""
    SELECT DISTINCT
        s.cliente_base,
        s.categoria,
        s.fabrica,
        s.ce_bri,
        s.familia,
        r.registro AS registro,
        s.tipo_processo,
        s.ip_processo,
        s.data_processo,
        s.arquivo_nome
    FROM sistema5_itens s
    LEFT JOIN certificados c
    ON UPPER(c.ce_bri) = UPPER(s.ce_bri)
    AND c.familia = s.familia
    LEFT JOIN ip_bri_familias m
    ON UPPER(m.ce_bri) = UPPER(s.ce_bri)
    AND m.familia = s.familia
    LEFT JOIN registros r
    ON UPPER(r.ce_bri) = UPPER(s.ce_bri)
    AND r.familia = s.familia
    WHERE s.ce_bri IS NOT NULL
    AND s.ce_bri != ''
    AND s.familia IS NOT NULL
    AND s.familia != ''
    AND c.ip_bri IS NULL
    AND m.ip_bri IS NULL
    ORDER BY s.cliente_base, s.fabrica, CAST(s.familia AS INTEGER), s.data_processo DESC
    """, conn)



def atualizar_ip_bri_manual_por_id(registro_id, novo_ip_bri, nova_observacao=""):
    registro_id = int(registro_id)
    novo_ip_bri = normalizar_ip_bri_manual(novo_ip_bri)
    nova_observacao = clean(nova_observacao)

    atual = cursor.execute("""
    SELECT ce_bri, familia
    FROM ip_bri_familias
    WHERE id = ?
    """, (registro_id,)).fetchone()

    if not atual:
        return False, "Cadastro manual não encontrado."

    ce_bri_atual, familia_atual = atual

    duplicado = cursor.execute("""
    SELECT ce_bri, familia
    FROM ip_bri_familias
    WHERE UPPER(ip_bri) = UPPER(?)
    AND id != ?
    LIMIT 1
    """, (novo_ip_bri, registro_id)).fetchone()

    if duplicado:
        return False, (
            f"Este IP-BRI já está cadastrado em {duplicado[0]} / família {duplicado[1]}. "
            f"Não é permitido repetir o mesmo IP-BRI em famílias diferentes."
        )

    cursor.execute("""
    UPDATE ip_bri_familias
    SET ip_bri = ?, observacao = ?, data_atualizacao = ?
    WHERE id = ?
    """, (
        novo_ip_bri,
        nova_observacao,
        datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        registro_id
    ))

    conn.commit()
    return True, f"Cadastro atualizado: {ce_bri_atual} / família {familia_atual} → {novo_ip_bri}."


def excluir_ip_bri_manual_por_id(registro_id):
    registro_id = int(registro_id)

    atual = cursor.execute("""
    SELECT ce_bri, familia, ip_bri
    FROM ip_bri_familias
    WHERE id = ?
    """, (registro_id,)).fetchone()

    if not atual:
        return False, "Cadastro manual não encontrado."

    cursor.execute("""
    DELETE FROM ip_bri_familias
    WHERE id = ?
    """, (registro_id,))

    conn.commit()
    return True, f"Cadastro excluído: {atual[0]} / família {atual[1]} / {atual[2]}."


def listar_ip_bri_manuais():
    return pd.read_sql_query("""
    SELECT
        id,
        ce_bri,
        familia,
        ip_bri,
        observacao,
        data_cadastro,
        data_atualizacao
    FROM ip_bri_familias
    ORDER BY ce_bri, CAST(familia AS INTEGER)
    """, conn)

# ==========================================
# PESQUISA AVANÇADA
# ==========================================

def pesquisar_por_ip_bri(ip_bri_busca):
    termo = clean(ip_bri_busca).upper()

    if not termo:
        return pd.DataFrame()

    termo_sem_prefixo = termo.replace("IP-BRI-", "")
    termo_com_prefixo = termo if termo.startswith("IP-BRI-") else f"IP-BRI-{termo}"

    return pd.read_sql_query("""
    SELECT DISTINCT
        c.ip_bri AS ip_bri,
        c.ce_bri AS ce_bri,
        c.familia AS fam,
        r.registro AS registro,
        r.fabrica AS fabrica,
        r.endereco_fabrica AS endereco,
        c.rev AS rev,
        c.produto AS produto,
        c.data_emissao AS data_emissao
    FROM certificados c
    LEFT JOIN registros r
    ON UPPER(r.ce_bri) = UPPER(c.ce_bri)
    AND r.familia = c.familia
    WHERE UPPER(c.ip_bri) LIKE ?
       OR REPLACE(UPPER(c.ip_bri), 'IP-BRI-', '') LIKE ?
    ORDER BY c.ip_bri, r.fabrica, r.registro
    """, conn, params=(f"{termo_com_prefixo}%", f"{termo_sem_prefixo}%"))


def pesquisar_por_registro(registro_busca):
    termo = clean(registro_busca)

    if not termo:
        return pd.DataFrame()

    return pd.read_sql_query("""
    SELECT
        r.registro AS registro,
        c.ip_bri AS ip_bri,
        c.ce_bri AS ce_bri,
        c.familia AS fam,
        r.fabrica AS fabrica,
        r.endereco_fabrica AS endereco,
        marcas.fabricante AS fabricante,
        c.rev AS rev,
        c.produto AS produto,
        c.data_emissao AS data_emissao
    FROM registros r
    LEFT JOIN certificados c
    ON UPPER(c.ce_bri) = UPPER(r.ce_bri)
    AND c.familia = r.familia
    LEFT JOIN (
        SELECT
            certificado_id,
            MIN(marca) AS fabricante
        FROM itens
        GROUP BY certificado_id
    ) marcas
    ON marcas.certificado_id = c.id
    WHERE r.registro LIKE ?
    ORDER BY r.registro, c.ip_bri, marcas.fabricante
    """, conn, params=(f"{termo}%",))


# ==========================================
# SISTEMA 5 - INCLUSÕES / MANUTENÇÕES
# ==========================================

def normalizar_categoria_sistema5(valor):
    valor = clean(valor).upper()
    if "NOVO" in valor:
        return "SISTEMA 5 NOVO PROJETO"
    if "PROPR" in valor:
        return "SISTEMA 5 PROPRIOS"
    return valor


def extrair_ip_processo(nome):
    texto = clean(nome).upper()
    match = re.search(r"\bIP[- ]?\d+[-/]\d+\b", texto)
    if match:
        return match.group(0).replace(" ", "-")
    return None


def extrair_data_processo(nome):
    texto = clean(nome)
    match = re.search(r"\b(\d{2})[-_.](\d{2})[-_.](\d{2,4})\b", texto)
    if not match:
        return None
    dia, mes, ano = match.groups()
    if len(ano) == 2:
        ano = "20" + ano
    return f"{ano}-{mes}-{dia}"


def detectar_tipo_processo(nome):
    texto = clean(nome).upper()
    if "MANUT" in texto:
        return "MANUTENCAO"
    if "RECERT" in texto:
        return "RECERTIFICACAO"
    if "INICIAL" in texto:
        return "INICIAL"
    if "INCLUS" in texto:
        return "INCLUSAO"
    return "OUTROS"


def get_or_create_cliente_sistema5(categoria, cliente_base):
    categoria = normalizar_categoria_sistema5(categoria)
    cliente_base = clean(cliente_base).upper()

    cursor.execute("""
    INSERT OR IGNORE INTO sistema5_clientes (categoria, cliente_base, data_cadastro)
    VALUES (?, ?, ?)
    """, (categoria, cliente_base, datetime.now().strftime("%d/%m/%Y %H:%M:%S")))
    conn.commit()

    return cursor.execute("""
    SELECT id FROM sistema5_clientes
    WHERE categoria = ? AND cliente_base = ?
    """, (categoria, cliente_base)).fetchone()[0]


def get_or_create_fabrica_sistema5(cliente_id, fabrica, ce_bri="", endereco_fabrica=""):
    fabrica = clean(fabrica).upper()
    ce_bri = clean(ce_bri).upper()
    endereco_fabrica = clean(endereco_fabrica)

    cursor.execute("""
    INSERT OR IGNORE INTO sistema5_fabricas (
        cliente_id, fabrica, ce_bri, endereco_fabrica, data_cadastro
    )
    VALUES (?, ?, ?, ?, ?)
    """, (
        cliente_id,
        fabrica,
        ce_bri,
        endereco_fabrica,
        datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    ))

    cursor.execute("""
    UPDATE sistema5_fabricas
    SET
        ce_bri = COALESCE(NULLIF(?, ''), ce_bri),
        endereco_fabrica = COALESCE(NULLIF(?, ''), endereco_fabrica)
    WHERE cliente_id = ? AND fabrica = ?
    """, (ce_bri, endereco_fabrica, cliente_id, fabrica))

    conn.commit()

    return cursor.execute("""
    SELECT id FROM sistema5_fabricas
    WHERE cliente_id = ? AND fabrica = ?
    """, (cliente_id, fabrica)).fetchone()[0]


def normalizar_coluna_excel_s5(valor):
    txt = clean(valor).upper()

    # Remove acentos para padronizar
    troca = {
        "Á": "A", "À": "A", "Â": "A", "Ã": "A",
        "É": "E", "Ê": "E",
        "Í": "I",
        "Ó": "O", "Ô": "O", "Õ": "O",
        "Ú": "U",
        "Ç": "C"
    }

    for antigo, novo in troca.items():
        txt = txt.replace(antigo, novo)

    txt = re.sub(r"\s+", " ", txt).strip()

    # Ignora colunas auxiliares que citam "modelo", mas não são referência.
    if "APLICAVEL SOMENTE" in txt or "QUANTIDADE" in txt or "NCM" in txt:
        return txt

    # Item
    if txt == "ITEM" or txt.startswith("ITEM "):
        return "ITEM"

    # Códigos
    if (
        "CODIGO DE BARRAS" in txt
        or "COD DE BARRAS" in txt
        or "COD. DE BARRAS" in txt
        or txt in ["CODIGO", "COD.", "COD"]
        or "BARRAS" in txt
    ):
        return "CODIGO"

    # Modelo / Referência
    # IMPORTANTE:
    # Não pode reconhecer qualquer texto com "MODELO",
    # porque existe coluna "Quantidade aplicável somente para o modelo 1b".
    if (
        txt.startswith("MODELO")
        or txt.startswith("REFERENCIA")
        or "MODELO / REFERENCIA" in txt
        or "MODELO/REFERENCIA" in txt
        or "DESIGNACAO COMERCIAL" in txt
    ):
        return "MODELO"

    # Marca
    if (
        txt == "MARCA"
        or txt.startswith("MARCA ")
        or "MARCA COMERCIALIZADA" in txt
        or "FABRICANTE" in txt
    ):
        return "MARCA"

    # Descrição técnica
    if (
        "DESCRICAO TECNICA" in txt
        or "DESCRICAO TECNICA DO MODELO" in txt
        or txt.startswith("DESCRICAO")
        or txt == "NOME"
        or "PROCESSO PRODUTIVO" in txt
    ):
        return "NOME"

    return txt



def valor_valido_item_s5(valor):
    texto = clean(valor)

    if not texto:
        return False

    if texto.lower() in ["nan", "none", "null", "-", "--"]:
        return False

    return True


def codigo_barras_valido_s5(valor):
    codigo = extrair_codigo_unico(valor)

    if not codigo:
        return None

    if len(codigo) < 8 or len(codigo) > 14:
        return None

    return codigo



def familia_desconsiderada_s5(nome_aba):
    texto = clean(nome_aba).upper()
    texto = texto.replace("Á", "A").replace("À", "A").replace("Â", "A").replace("Ã", "A")
    texto = texto.replace("É", "E").replace("Ê", "E")
    texto = texto.replace("Í", "I")
    texto = texto.replace("Ó", "O").replace("Ô", "O").replace("Õ", "O")
    texto = texto.replace("Ú", "U")
    texto = texto.replace("Ç", "C")

    return "DESCONSIDERADO" in texto or "DESCONSIDERADA" in texto



def limpar_itens_desconsiderados_sistema5():
    """
    Remove itens antigos que já foram cadastrados no Sistema 5
    em famílias/abas marcadas como DESCONSIDERADO ou DESCONSIDERADA.
    """

    try:
        qtd_antes = cursor.execute("""
        SELECT COUNT(*)
        FROM sistema5_itens
        WHERE UPPER(familia) LIKE '%DESCONSIDERADO%'
        OR UPPER(familia) LIKE '%DESCONSIDERADA%'
        """).fetchone()[0]

        cursor.execute("""
        DELETE FROM sistema5_itens
        WHERE UPPER(familia) LIKE '%DESCONSIDERADO%'
        OR UPPER(familia) LIKE '%DESCONSIDERADA%'
        """)

        conn.commit()

        return qtd_antes

    except Exception:
        return 0



def parece_codigo_barras_s5(valor):
    codigo = re.sub(r"\D", "", clean(valor))
    return bool(codigo and len(codigo) >= 8 and len(codigo) <= 14)


def parece_marca_s5(valor):
    texto = clean(valor)
    if not texto or texto.lower() in ["nan", "none", "null"]:
        return False
    # Marca normalmente tem letras; código de barras normalmente só números.
    return bool(re.search(r"[A-Za-zÀ-ÿ]", texto))


def corrigir_marca_codigo_invertidos_s5(marca, codigo_raw):
    """
    Alguns memoriais vêm com o cabeçalho:
    Marca Comercializada | Código de Barras
    mas os dados vêm invertidos:
    código na coluna da marca e marca na coluna do código.

    Se detectar isso, troca automaticamente.
    """
    marca_limpa = clean(marca)
    codigo_limpo = clean(codigo_raw)

    if parece_codigo_barras_s5(marca_limpa) and parece_marca_s5(codigo_limpo):
        return codigo_limpo, marca_limpa

    return marca_limpa, codigo_limpo


def ler_excel_inclusao_sistema5(uploaded_file):
    """
    Leitor específico do Sistema 5.

    Estrutura esperada:
    - O Excel pode ter várias abas.
    - Cada aba representa uma família.
    - Dentro de cada aba existem as colunas:
      ITEM, MODELO / REFERÊNCIA, MARCA COMERCIALIZADA,
      CÓDIGO DE BARRAS, DESCRIÇÃO TÉCNICA.

    Regras:
    - Só considera linha real se tiver MODELO + MARCA + CÓDIGO + DESCRIÇÃO.
    - Ignora observações, assinaturas, local/data, rodapé e linhas soltas.
    """

    try:
        excel = pd.ExcelFile(uploaded_file)
    except Exception:
        return pd.DataFrame()

    todos_itens = []

    for aba in excel.sheet_names:
        # Se a aba/família estiver marcada como DESCONSIDERADO,
        # nenhum item desta família será cadastrado.
        if familia_desconsiderada_s5(aba):
            continue

        try:
            bruto = pd.read_excel(
                uploaded_file,
                sheet_name=aba,
                header=None,
                dtype=object
            )
        except Exception:
            continue

        if bruto.empty:
            continue

        header_idx = None
        colunas = {}

        for i, row in bruto.iterrows():
            temp = {}

            for idx, valor in enumerate(row.values):
                nome = normalizar_coluna_excel_s5(valor)

                if nome == "ITEM":
                    temp["ITEM"] = idx
                elif nome == "MODELO":
                    temp["MODELO"] = idx
                elif nome == "MARCA":
                    temp["MARCA"] = idx
                elif nome == "CODIGO":
                    temp["CODIGO"] = idx
                elif nome == "NOME":
                    temp["NOME"] = idx

            # Cabeçalho real precisa ter os campos centrais.
            if all(campo in temp for campo in ["MODELO", "MARCA", "CODIGO", "NOME"]):
                header_idx = i
                colunas = temp
                break

        if header_idx is None:
            continue

        for j in range(header_idx + 1, len(bruto)):
            row = bruto.iloc[j]

            def valor_coluna(nome_coluna):
                idx = colunas.get(nome_coluna)
                if idx is None or idx >= len(row):
                    return ""
                return clean(row.iloc[idx])

            item = valor_coluna("ITEM")
            modelo = valor_coluna("MODELO")
            marca = valor_coluna("MARCA")
            codigo_raw = valor_coluna("CODIGO")
            nome = valor_coluna("NOME")

            # Corrige automaticamente quando o Excel vem com MARCA e CÓDIGO invertidos.
            marca, codigo_raw = corrigir_marca_codigo_invertidos_s5(marca, codigo_raw)

            codigo = codigo_barras_valido_s5(codigo_raw)

            linha_texto = " ".join([modelo, marca, codigo_raw, nome]).upper()

            palavras_ignorar = [
                "OBS",
                "OBSERVAÇÃO",
                "OBSERVACAO",
                "LOCAL E DATA",
                "PLACE AND DATE",
                "ASSINATURA",
                "SIGNATURE",
                "RESPONSÁVEL",
                "RESPONSAVEL",
                "APROVAÇÃO",
                "APROVACAO"
            ]

            if any(p in linha_texto for p in palavras_ignorar):
                continue

            # Linha real precisa ter os 4 campos principais válidos.
            if not valor_valido_item_s5(modelo):
                continue

            if not valor_valido_item_s5(marca):
                continue

            if not valor_valido_item_s5(nome):
                continue

            # Código precisa ser código de barras válido, com 8 a 14 dígitos.
            if not codigo:
                continue

            # Evita linhas de complemento/rodapé que não são itens reais.
            if marca.upper().strip() in ["DIRETA", "INDIRETA"]:
                continue

            if modelo.upper().strip() in ["DIRETA", "INDIRETA"]:
                continue

            if nome.upper().strip() in ["DIRETA", "INDIRETA"]:
                continue

            # Evita cabeçalho repetido
            if normalizar_coluna_excel_s5(modelo) == "MODELO":
                continue

            if normalizar_coluna_excel_s5(marca) == "MARCA":
                continue

            todos_itens.append({
                "FAMILIA": normalizar_familia_aba_s5(aba),
                "ITEM": item,
                "MARCA": marca,
                "MODELO": modelo,
                "NOME": nome,
                "CODIGO": codigo
            })

    if not todos_itens:
        return pd.DataFrame()

    df_final = pd.DataFrame(todos_itens)

    df_final = df_final.drop_duplicates(
        subset=["FAMILIA", "MODELO", "CODIGO"],
        keep="first"
    )

    return df_final


def ip_processo_ja_existe_sistema5(ip_processo, cliente_base="", fabrica=""):
    """
    Retorna os processos já salvos com o mesmo IP.
    Filtra por cliente_base e fabrica se informados.
    """
    ip = clean(ip_processo).upper()

    where = ["UPPER(a.ip_processo) = ?"]
    params = [ip]

    if clean(cliente_base):
        where.append("UPPER(c.cliente_base) = UPPER(?)")
        params.append(clean(cliente_base))

    if clean(fabrica):
        where.append("UPPER(f.fabrica) = UPPER(?)")
        params.append(clean(fabrica))

    sql = f"""
    SELECT
        a.id AS arquivo_id,
        c.categoria,
        c.cliente_base,
        f.fabrica,
        a.tipo_processo,
        a.ip_processo,
        a.data_processo,
        a.arquivo_nome,
        COUNT(i.id) AS qtd_itens
    FROM sistema5_arquivos a
    INNER JOIN sistema5_clientes c ON c.id = a.cliente_id
    INNER JOIN sistema5_fabricas f ON f.id = a.fabrica_id
    LEFT JOIN sistema5_itens i ON i.arquivo_id = a.id
    WHERE {" AND ".join(where)}
    GROUP BY a.id
    ORDER BY a.id DESC
    """

    return pd.read_sql_query(sql, conn, params=tuple(params))


def deletar_processo_sistema5(arquivo_id):
    """
    Remove um processo do Sistema 5 e todos os seus itens.
    Retorna o número de itens deletados.
    """
    arquivo_id = int(arquivo_id)

    qtd = cursor.execute(
        "SELECT COUNT(*) FROM sistema5_itens WHERE arquivo_id = ?",
        (arquivo_id,)
    ).fetchone()[0]

    cursor.execute("DELETE FROM sistema5_itens WHERE arquivo_id = ?", (arquivo_id,))
    cursor.execute("DELETE FROM sistema5_arquivos WHERE id = ?", (arquivo_id,))
    conn.commit()

    return qtd


def salvar_inclusao_sistema5(categoria, cliente_base, fabrica, ce_bri, endereco_fabrica, tipo_processo, ip_processo, data_processo, arquivo_nome, df_itens):
    categoria = normalizar_categoria_sistema5(categoria)
    cliente_base = clean(cliente_base).upper()
    fabrica = clean(fabrica).upper()
    ce_bri = clean(ce_bri).upper()
    endereco_fabrica = clean(endereco_fabrica)
    tipo_processo = clean(tipo_processo).upper()
    ip_processo = clean(ip_processo).upper()

    cliente_id = get_or_create_cliente_sistema5(categoria, cliente_base)
    fabrica_id = get_or_create_fabrica_sistema5(cliente_id, fabrica, ce_bri, endereco_fabrica)

    cursor.execute("""
    INSERT INTO sistema5_arquivos (
        cliente_id, fabrica_id, tipo_processo, ip_processo, data_processo, arquivo_nome, data_upload
    )
    VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (
        cliente_id,
        fabrica_id,
        tipo_processo,
        ip_processo,
        data_processo,
        arquivo_nome,
        datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    ))
    conn.commit()
    arquivo_id = cursor.lastrowid

    total = 0

    for _, r in df_itens.iterrows():
        cursor.execute("""
        INSERT INTO sistema5_itens (
            arquivo_id, cliente_base, categoria, fabrica, ce_bri, endereco_fabrica,
            tipo_processo, ip_processo, data_processo, familia, item, marca, modelo, nome, codigo,
            arquivo_nome, data_upload
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            arquivo_id,
            cliente_base,
            categoria,
            fabrica,
            ce_bri,
            endereco_fabrica,
            tipo_processo,
            ip_processo,
            data_processo,
            clean(r.get("FAMILIA", "")),
            clean(r.get("ITEM", "")),
            clean(r.get("MARCA", "")),
            clean(r.get("MODELO", "")),
            clean(r.get("NOME", "")),
            clean(r.get("CODIGO", "")),
            arquivo_nome,
            datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        ))
        total += 1

    conn.commit()
    return total


def buscar_item_sistema5_confirmacao(valor_ref):
    ref = clean(valor_ref)

    if not ref:
        return None

    # Busca candidatos, mas só aceita se a referência inicial for exatamente igual.
    resultado = pd.read_sql_query("""
    SELECT
        marca AS MARCA,
        modelo AS MODELO,
        nome AS NOME,
        codigo AS CODIGO,
        ip_processo AS IP_PROCESSO,
        ce_bri AS CE_BRI,
        NULL AS REGISTRO,
        endereco_fabrica AS ENDERECO,
        fabrica AS FABRICA,
        tipo_processo AS TIPO_PROCESSO,
        data_processo AS DATA_PROCESSO,
        familia AS FAMILIA,
        arquivo_nome AS ARQUIVO_ORIGEM
    FROM sistema5_itens
    WHERE UPPER(TRIM(modelo)) = UPPER(TRIM(?))
       OR UPPER(TRIM(modelo)) LIKE UPPER(TRIM(?) || ' -%')
       OR UPPER(TRIM(modelo)) LIKE UPPER(TRIM(?) || '%')
    ORDER BY COALESCE(data_processo, '1900-01-01') DESC, id DESC
    LIMIT 100
    """, conn, params=(ref, ref, ref))

    resultado = filtrar_resultado_por_referencia_exata(resultado, ref)

    if resultado.empty:
        return None

    item = resultado.iloc[0].to_dict()

    ce_bri = clean(item.get("CE_BRI"))
    familia = clean(item.get("FAMILIA"))

    # IP do processo fica separado em IP_PROCESSO.
    # IP_BRI tenta puxar o IP-BRI oficial pelo CE-BRI + família.
    # Se não existir certificado oficial ainda, usa a família encontrada.
    ip_bri_oficial = None

    if ce_bri and familia:
        oficial = pd.read_sql_query("""
        SELECT ip_bri
        FROM certificados
        WHERE UPPER(ce_bri) = UPPER(?)
        AND familia = ?
        ORDER BY rev DESC, id DESC
        LIMIT 1
        """, conn, params=(ce_bri, familia))

        if not oficial.empty:
            ip_bri_oficial = clean(oficial.iloc[0]["ip_bri"])

    if ip_bri_oficial:
        item["IP_BRI"] = ip_bri_oficial
    else:
        ip_bri_manual = buscar_ip_bri_manual_familia(ce_bri, familia)
        if ip_bri_manual:
            item["IP_BRI"] = ip_bri_manual
        else:
            item["IP_BRI"] = familia

    if ce_bri and familia:
        registro_complementar = pd.read_sql_query("""
        SELECT registro
        FROM registros
        WHERE UPPER(ce_bri) = UPPER(?)
        AND familia = ?
        ORDER BY id DESC
        LIMIT 1
        """, conn, params=(ce_bri, familia))

        if not registro_complementar.empty:
            item["REGISTRO"] = registro_complementar.iloc[0]["registro"]

    return item


def listar_sistema5_resumo():
    return pd.read_sql_query("""
    SELECT
        c.categoria,
        c.cliente_base,
        f.fabrica,
        f.ce_bri,
        f.endereco_fabrica,
        a.tipo_processo,
        a.ip_processo,
        a.data_processo,
        a.arquivo_nome,
        COUNT(i.id) AS qtd_itens
    FROM sistema5_clientes c
    LEFT JOIN sistema5_fabricas f ON f.cliente_id = c.id
    LEFT JOIN sistema5_arquivos a ON a.fabrica_id = f.id
    LEFT JOIN sistema5_itens i ON i.arquivo_id = a.id
    GROUP BY c.categoria, c.cliente_base, f.fabrica, f.ce_bri, f.endereco_fabrica,
             a.tipo_processo, a.ip_processo, a.data_processo, a.arquivo_nome
    ORDER BY c.categoria, c.cliente_base, f.fabrica, a.data_processo DESC
    """, conn)


def buscar_dados_fabrica_existente(cliente_base="", fabrica="", ce_bri=""):
    cliente_base = clean(cliente_base).upper()
    fabrica = clean(fabrica)
    ce_bri = clean(ce_bri).upper()

    where = []
    params = []

    if cliente_base:
        where.append("UPPER(cliente_base) = UPPER(?)")
        params.append(cliente_base)

    if fabrica:
        where.append("UPPER(fabrica) = UPPER(?)")
        params.append(fabrica)

    if ce_bri:
        where.append("UPPER(ce_bri) = UPPER(?)")
        params.append(ce_bri)

    if not where:
        return None

    sql = f"""
    SELECT
        cliente_base,
        fabrica,
        ce_bri,
        endereco_fabrica
    FROM registros
    WHERE {" AND ".join(where)}
    ORDER BY id DESC
    LIMIT 1
    """

    resultado = pd.read_sql_query(sql, conn, params=tuple(params))

    if resultado.empty:
        return None

    return resultado.iloc[0].to_dict()


def listar_fabricas_existentes_para_sistema5(cliente_base=""):
    cliente_base = clean(cliente_base).upper()

    if cliente_base:
        return pd.read_sql_query("""
        SELECT DISTINCT
            fabrica,
            ce_bri,
            endereco_fabrica
        FROM registros
        WHERE UPPER(cliente_base) = UPPER(?)
        AND fabrica IS NOT NULL
        AND fabrica != ''
        ORDER BY fabrica
        """, conn, params=(cliente_base,))

    return pd.read_sql_query("""
    SELECT DISTINCT
        cliente_base,
        fabrica,
        ce_bri,
        endereco_fabrica
    FROM registros
    WHERE fabrica IS NOT NULL
    AND fabrica != ''
    ORDER BY cliente_base, fabrica
    """, conn)


def extrair_codigo_fabrica_nome(nome):
    texto = clean(nome).upper()
    match = re.search(r"\bF\s*0*(\d{1,3})\b", texto)
    if match:
        numero = int(match.group(1))
        return f"F{numero:02d}"
    return ""


def normalizar_familia_aba_s5(nome_aba):
    texto = clean(nome_aba).upper()
    numeros = re.findall(r"\d+", texto)
    if numeros:
        return str(int(numeros[0]))
    return texto


def buscar_fabrica_por_codigo_s5(cliente_base, codigo_fabrica):
    cliente_base = clean(cliente_base).upper()
    codigo_fabrica = clean(codigo_fabrica).upper()

    if not cliente_base or not codigo_fabrica:
        return None

    numero = re.sub(r"\D", "", codigo_fabrica)
    if numero:
        numero_int = str(int(numero))
        codigo_padrao = f"F{int(numero):02d}"
    else:
        numero_int = ""
        codigo_padrao = codigo_fabrica

    resultado = pd.read_sql_query("""
    SELECT
        cliente_base,
        fabrica,
        ce_bri,
        endereco_fabrica
    FROM registros
    WHERE UPPER(cliente_base) = UPPER(?)
    AND (
        UPPER(fabrica) LIKE ?
        OR UPPER(fabrica) LIKE ?
        OR UPPER(fabrica) LIKE ?
    )
    ORDER BY id DESC
    LIMIT 1
    """, conn, params=(
        cliente_base,
        f"%{codigo_padrao}%",
        f"%F{numero_int}%" if numero_int else f"%{codigo_fabrica}%",
        f"%FÁBRICA {numero_int}%" if numero_int else f"%{codigo_fabrica}%"
    ))

    if resultado.empty:
        return None

    return resultado.iloc[0].to_dict()


def listar_processos_sistema5():
    return pd.read_sql_query("""
    SELECT
        a.id AS arquivo_id,
        c.categoria,
        c.cliente_base,
        f.fabrica,
        f.ce_bri,
        f.endereco_fabrica,
        a.tipo_processo,
        a.ip_processo,
        a.data_processo,
        a.arquivo_nome,
        COUNT(i.id) AS qtd_itens
    FROM sistema5_arquivos a
    INNER JOIN sistema5_clientes c
    ON c.id = a.cliente_id
    INNER JOIN sistema5_fabricas f
    ON f.id = a.fabrica_id
    LEFT JOIN sistema5_itens i
    ON i.arquivo_id = a.id
    GROUP BY
        a.id,
        c.categoria,
        c.cliente_base,
        f.fabrica,
        f.ce_bri,
        f.endereco_fabrica,
        a.tipo_processo,
        a.ip_processo,
        a.data_processo,
        a.arquivo_nome
    ORDER BY
        c.categoria,
        c.cliente_base,
        f.fabrica,
        a.data_processo DESC,
        a.id DESC
    """, conn)


def buscar_itens_processo_sistema5(arquivo_id, termo=""):
    termo = clean(termo)

    if termo:
        return pd.read_sql_query("""
        SELECT
            familia,
            item,
            marca,
            modelo,
            nome,
            codigo,
            ce_bri,
            endereco_fabrica,
            tipo_processo,
            ip_processo,
            data_processo,
            arquivo_nome
        FROM sistema5_itens
        WHERE arquivo_id = ?
        AND (
            UPPER(modelo) LIKE UPPER(?)
            OR UPPER(marca) LIKE UPPER(?)
            OR UPPER(nome) LIKE UPPER(?)
            OR codigo LIKE ?
        )
        ORDER BY CAST(familia AS INTEGER), CAST(item AS INTEGER), modelo
        """, conn, params=(
            arquivo_id,
            f"{termo}%",
            f"%{termo}%",
            f"%{termo}%",
            f"%{termo}%"
        ))

    return pd.read_sql_query("""
    SELECT
        familia,
        item,
        marca,
        modelo,
        nome,
        codigo,
        ce_bri,
        endereco_fabrica,
        tipo_processo,
        ip_processo,
        data_processo,
        arquivo_nome
    FROM sistema5_itens
    WHERE arquivo_id = ?
    ORDER BY CAST(familia AS INTEGER), CAST(item AS INTEGER), modelo
    """, conn, params=(arquivo_id,))


# Limpeza automática de itens antigos do Sistema 5 marcados como DESCONSIDERADO
try:
    limpar_itens_desconsiderados_sistema5()
except Exception:
    pass


def pesquisar_global_sistema5(termo):
    termo = clean(termo)

    if not termo:
        return pd.DataFrame()

    return pd.read_sql_query("""
    SELECT
        cliente_base,
        categoria,
        fabrica,
        ce_bri,
        endereco_fabrica,
        tipo_processo,
        ip_processo,
        data_processo,
        familia,
        item,
        marca,
        modelo,
        nome,
        codigo,
        arquivo_nome,
        data_upload
    FROM sistema5_itens
    WHERE
        UPPER(modelo) LIKE UPPER(?)
        OR codigo LIKE ?
        OR UPPER(marca) LIKE UPPER(?)
        OR UPPER(nome) LIKE UPPER(?)
        OR UPPER(ip_processo) LIKE UPPER(?)
        OR UPPER(ce_bri) LIKE UPPER(?)
        OR UPPER(fabrica) LIKE UPPER(?)
        OR UPPER(cliente_base) LIKE UPPER(?)
        OR UPPER(arquivo_nome) LIKE UPPER(?)
    ORDER BY
        COALESCE(data_processo, '1900-01-01') DESC,
        id DESC
    """, conn, params=(
        f"{termo}%",
        f"%{termo}%",
        f"%{termo}%",
        f"%{termo}%",
        f"%{termo}%",
        f"%{termo}%",
        f"%{termo}%",
        f"%{termo}%",
        f"%{termo}%"
    ))



# ==========================================
# PAINEL DE COBERTURA
# ==========================================

def cobertura_resumo_geral():
    dados = {}

    try:
        dados["clientes_registros"] = cursor.execute("""
        SELECT COUNT(DISTINCT cliente_base)
        FROM registros
        WHERE cliente_base IS NOT NULL AND cliente_base != ''
        """).fetchone()[0]
    except Exception:
        dados["clientes_registros"] = 0

    try:
        dados["fabricas_registros"] = cursor.execute("""
        SELECT COUNT(DISTINCT cliente_base || '|' || fabrica)
        FROM registros
        WHERE fabrica IS NOT NULL AND fabrica != ''
        """).fetchone()[0]
    except Exception:
        dados["fabricas_registros"] = 0

    try:
        dados["familias_registros"] = cursor.execute("""
        SELECT COUNT(DISTINCT ce_bri || '|' || familia)
        FROM registros
        WHERE ce_bri IS NOT NULL AND ce_bri != ''
        AND familia IS NOT NULL AND familia != ''
        """).fetchone()[0]
    except Exception:
        dados["familias_registros"] = 0

    try:
        dados["familias_sistema5"] = cursor.execute("""
        SELECT COUNT(DISTINCT ce_bri || '|' || familia)
        FROM sistema5_itens
        WHERE ce_bri IS NOT NULL AND ce_bri != ''
        AND familia IS NOT NULL AND familia != ''
        """).fetchone()[0]
    except Exception:
        dados["familias_sistema5"] = 0

    try:
        dados["itens_sistema5"] = cursor.execute("""
        SELECT COUNT(*)
        FROM sistema5_itens
        """).fetchone()[0]
    except Exception:
        dados["itens_sistema5"] = 0

    try:
        dados["pendentes_ip_bri"] = len(listar_ip_bri_pendentes_sistema5())
    except Exception:
        dados["pendentes_ip_bri"] = 0

    return dados


def cobertura_fabricas():
    return pd.read_sql_query("""
    WITH chaves AS (
        SELECT DISTINCT cliente_base, fabrica, ce_bri
        FROM registros
        WHERE fabrica IS NOT NULL AND fabrica != ''

        UNION

        SELECT DISTINCT cliente_base, fabrica, ce_bri
        FROM sistema5_itens
        WHERE fabrica IS NOT NULL AND fabrica != ''
    ),
    familias_reg AS (
        SELECT
            cliente_base,
            fabrica,
            ce_bri,
            COUNT(DISTINCT familia) AS familias_registros
        FROM registros
        WHERE fabrica IS NOT NULL AND fabrica != ''
        GROUP BY cliente_base, fabrica, ce_bri
    ),
    familias_s5 AS (
        SELECT
            cliente_base,
            fabrica,
            ce_bri,
            COUNT(DISTINCT familia) AS familias_sistema5,
            COUNT(*) AS itens_sistema5
        FROM sistema5_itens
        WHERE fabrica IS NOT NULL AND fabrica != ''
        GROUP BY cliente_base, fabrica, ce_bri
    ),
    pendentes AS (
        SELECT
            s.cliente_base,
            s.fabrica,
            s.ce_bri,
            COUNT(DISTINCT s.familia) AS familias_pendentes_ip_bri
        FROM sistema5_itens s
        LEFT JOIN certificados c
        ON UPPER(c.ce_bri) = UPPER(s.ce_bri)
        AND c.familia = s.familia
        LEFT JOIN ip_bri_familias m
        ON UPPER(m.ce_bri) = UPPER(s.ce_bri)
        AND m.familia = s.familia
        WHERE s.ce_bri IS NOT NULL
        AND s.ce_bri != ''
        AND s.familia IS NOT NULL
        AND s.familia != ''
        AND c.ip_bri IS NULL
        AND m.ip_bri IS NULL
        GROUP BY s.cliente_base, s.fabrica, s.ce_bri
    )
    SELECT
        chaves.cliente_base,
        chaves.fabrica,
        chaves.ce_bri,
        COALESCE(r.familias_registros, 0) AS familias_registros,
        COALESCE(s.familias_sistema5, 0) AS familias_sistema5,
        COALESCE(s.itens_sistema5, 0) AS itens_sistema5,
        COALESCE(p.familias_pendentes_ip_bri, 0) AS familias_pendentes_ip_bri
    FROM chaves
    LEFT JOIN familias_reg r
    ON UPPER(r.cliente_base) = UPPER(chaves.cliente_base)
    AND UPPER(r.fabrica) = UPPER(chaves.fabrica)
    AND UPPER(r.ce_bri) = UPPER(chaves.ce_bri)
    LEFT JOIN familias_s5 s
    ON UPPER(s.cliente_base) = UPPER(chaves.cliente_base)
    AND UPPER(s.fabrica) = UPPER(chaves.fabrica)
    AND UPPER(s.ce_bri) = UPPER(chaves.ce_bri)
    LEFT JOIN pendentes p
    ON UPPER(p.cliente_base) = UPPER(chaves.cliente_base)
    AND UPPER(p.fabrica) = UPPER(chaves.fabrica)
    AND UPPER(p.ce_bri) = UPPER(chaves.ce_bri)
    ORDER BY chaves.cliente_base, chaves.fabrica
    """, conn)

def cobertura_familias_detalhada(cliente_base="", fabrica="", ce_bri=""):
    where = []
    params = []

    if clean(cliente_base):
        where.append("UPPER(base.cliente_base) = UPPER(?)")
        params.append(clean(cliente_base))

    if clean(fabrica):
        where.append("UPPER(base.fabrica) = UPPER(?)")
        params.append(clean(fabrica))

    if clean(ce_bri):
        where.append("UPPER(base.ce_bri) = UPPER(?)")
        params.append(clean(ce_bri))

    filtro = ""
    if where:
        filtro = "WHERE " + " AND ".join(where)

    sql = f"""
    WITH base AS (
        SELECT
            cliente_base,
            fabrica,
            ce_bri,
            familia
        FROM registros
        WHERE familia IS NOT NULL AND familia != ''

        UNION

        SELECT
            cliente_base,
            fabrica,
            ce_bri,
            familia
        FROM sistema5_itens
        WHERE familia IS NOT NULL AND familia != ''
    ),
    itens_s5 AS (
        SELECT
            cliente_base,
            fabrica,
            ce_bri,
            familia,
            COUNT(*) AS qtd_itens_sistema5,
            GROUP_CONCAT(DISTINCT ip_processo) AS ips_processos
        FROM sistema5_itens
        GROUP BY cliente_base, fabrica, ce_bri, familia
    ),
    regs AS (
        SELECT
            cliente_base,
            fabrica,
            ce_bri,
            familia,
            MAX(registro) AS registro
        FROM registros
        GROUP BY cliente_base, fabrica, ce_bri, familia
    ),
    certs AS (
        SELECT
            ce_bri,
            familia,
            MAX(ip_bri) AS ip_bri_oficial,
            MAX(rev) AS rev
        FROM certificados
        GROUP BY ce_bri, familia
    ),
    manuais AS (
        SELECT
            ce_bri,
            familia,
            MAX(ip_bri) AS ip_bri_manual
        FROM ip_bri_familias
        GROUP BY ce_bri, familia
    )
    SELECT
        base.cliente_base,
        base.fabrica,
        base.ce_bri,
        base.familia,
        COALESCE(itens_s5.qtd_itens_sistema5, 0) AS qtd_itens_sistema5,
        COALESCE(regs.registro, '') AS registro,
        COALESCE(certs.ip_bri_oficial, '') AS ip_bri_oficial,
        COALESCE(manuais.ip_bri_manual, '') AS ip_bri_manual,
        CASE
            WHEN certs.ip_bri_oficial IS NOT NULL AND certs.ip_bri_oficial != '' THEN 'OFICIAL'
            WHEN manuais.ip_bri_manual IS NOT NULL AND manuais.ip_bri_manual != '' THEN 'MANUAL'
            ELSE 'PENDENTE'
        END AS status_ip_bri,
        COALESCE(certs.rev, '') AS rev,
        COALESCE(itens_s5.ips_processos, '') AS ips_processos
    FROM base
    LEFT JOIN itens_s5
    ON UPPER(itens_s5.cliente_base) = UPPER(base.cliente_base)
    AND UPPER(itens_s5.fabrica) = UPPER(base.fabrica)
    AND UPPER(itens_s5.ce_bri) = UPPER(base.ce_bri)
    AND itens_s5.familia = base.familia
    LEFT JOIN regs
    ON UPPER(regs.cliente_base) = UPPER(base.cliente_base)
    AND UPPER(regs.fabrica) = UPPER(base.fabrica)
    AND UPPER(regs.ce_bri) = UPPER(base.ce_bri)
    AND regs.familia = base.familia
    LEFT JOIN certs
    ON UPPER(certs.ce_bri) = UPPER(base.ce_bri)
    AND certs.familia = base.familia
    LEFT JOIN manuais
    ON UPPER(manuais.ce_bri) = UPPER(base.ce_bri)
    AND manuais.familia = base.familia
    {filtro}
    ORDER BY base.cliente_base, base.fabrica, CAST(base.familia AS INTEGER)
    """

    return pd.read_sql_query(sql, conn, params=tuple(params))


def cobertura_itens_por_familia(cliente_base, fabrica, ce_bri, familia):
    return pd.read_sql_query("""
    SELECT
        familia,
        item,
        marca,
        modelo,
        nome,
        codigo,
        tipo_processo,
        ip_processo,
        data_processo,
        arquivo_nome
    FROM sistema5_itens
    WHERE UPPER(cliente_base) = UPPER(?)
    AND UPPER(fabrica) = UPPER(?)
    AND UPPER(ce_bri) = UPPER(?)
    AND familia = ?
    ORDER BY CAST(item AS INTEGER), modelo
    """, conn, params=(
        clean(cliente_base),
        clean(fabrica),
        clean(ce_bri),
        normalizar_familia(familia)
    ))

# ==========================================
# TABS
# ==========================================

aviso_backup_diario()

try:
    st.sidebar.header("Backup Geral")

    st.sidebar.caption("Backup centralizado: baixar e restaurar tudo pelo mesmo local.")

    st.sidebar.download_button(
        "⬇️ BAIXAR BACKUP GERAL ZIP",
        gerar_backup_geral_zip(),
        "backup_geral_c_xml_br_engine.zip",
        "application/zip",
        key="backup_geral_sidebar"
    )

    st.sidebar.divider()

    st.sidebar.subheader("Restaurar backup")

    backup_geral_upload = st.sidebar.file_uploader(
        "Enviar certificados.db",
        type=["db"],
        key="backup_geral_upload_sidebar"
    )

    if backup_geral_upload:
        st.sidebar.warning("⚠️ Isso irá **substituir completamente** o banco atual.")
        confirmar_sidebar = st.sidebar.checkbox(
            "Confirmo que quero substituir o banco",
            key="confirmar_restore_sidebar"
        )
        if confirmar_sidebar:
            with open(DB_PATH, "wb") as f:
                f.write(backup_geral_upload.read())
            st.sidebar.success("Backup restaurado com sucesso.")
            st.sidebar.warning("Recarregue o app para aplicar o banco restaurado.")

except Exception as e:
    st.sidebar.warning(f"Backup geral indisponível: {e}")

aba0, aba1, aba2, aba3, aba4, aba5, aba6, aba7, aba8 = st.tabs([
    "🏠 Início",
    "PDF → XML",
    "Banco de Certificados",
    "Registros",
    "Preenchimento de Confirmação",
    "Pesquisa Avançada",
    "Sistema 5",
    "IP-BRI Pendentes",
    "Painel de Cobertura"
])

# ==========================================
# ABA 0 - INÍCIO / COMO USAR
# ==========================================

with aba0:
    st.title("C XML BR Engine")
    st.caption("Sistema de gestão de certificados IP-BRI, geração de XML e preenchimento automático de confirmações.")

    st.divider()

    # Status rápido do banco
    try:
        total_certs_home = cursor.execute("SELECT COUNT(*) FROM certificados").fetchone()[0]
        total_itens_home = cursor.execute("SELECT COUNT(*) FROM itens").fetchone()[0]
        total_regs_home  = cursor.execute("SELECT COUNT(*) FROM registros").fetchone()[0]
        total_s5_home    = cursor.execute("SELECT COUNT(*) FROM sistema5_itens").fetchone()[0]

        st.subheader("📊 Status do banco")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Certificados", total_certs_home)
        c2.metric("Itens", total_itens_home)
        c3.metric("Registros", total_regs_home)
        c4.metric("Itens Sistema 5", total_s5_home)

        if total_certs_home == 0:
            st.warning("⚠️ O banco está vazio. Siga o fluxo abaixo para começar.")
        else:
            st.success("Banco com dados. Pronto para uso ✅")
    except Exception:
        pass

    st.divider()

    st.subheader("🗺️ Fluxo de uso recomendado")
    st.caption("Siga esta ordem ao usar o sistema pela primeira vez ou ao cadastrar um novo cliente.")

    col_p1, col_seta1, col_p2, col_seta2, col_p3, col_seta3, col_p4 = st.columns([3, 0.5, 3, 0.5, 3, 0.5, 3])

    with col_p1:
        st.markdown("""
**Passo 1 — PDF → XML**

Envie o PDF do certificado IP-BRI.

O sistema extrai os itens, gera o XML e salva automaticamente no banco.

➡️ Use a aba **PDF → XML**
""")

    with col_seta1:
        st.markdown("<br><br><br>▶", unsafe_allow_html=True)

    with col_p2:
        st.markdown("""
**Passo 2 — Registros**

Envie o Excel de registros do cliente (fábricas, CE-BRI, números de registro).

Cada aba do Excel é uma fábrica.

➡️ Use a aba **Registros**
""")

    with col_seta2:
        st.markdown("<br><br><br>▶", unsafe_allow_html=True)

    with col_p3:
        st.markdown("""
**Passo 3 — Preenchimento**

Envie o Excel de confirmação com células verdes (referências) e amarelas (campos a preencher).

O sistema cruza com o banco e preenche automaticamente.

➡️ Use a aba **Preenchimento de Confirmação**
""")

    with col_seta3:
        st.markdown("<br><br><br>▶", unsafe_allow_html=True)

    with col_p4:
        st.markdown("""
**Passo 4 — Pesquisa / Histórico**

Consulte cruzamentos entre IP-BRI, registro e fábrica, ou veja o histórico de alterações de um certificado.

➡️ Use a aba **Pesquisa Avançada** ou **Banco de Certificados**
""")

    st.divider()

    st.subheader("📋 Referência rápida por aba")

    with st.expander("PDF → XML — O que faz e quando usar"):
        st.markdown("""
- Lê o PDF do certificado e extrai os itens (MARCA, MODELO, NOME, CÓDIGO DE BARRAS)
- Gera dois XMLs: um com vírgula e um com ponto no final do MODELO
- Salva automaticamente o certificado no banco (IP-BRI, CE-BRI, FAMÍLIA, REV)
- Se já existir um certificado com o mesmo IP-BRI e REV maior no banco, o envio é **ignorado**
- **Atenção:** se o PDF não tiver IP-BRI identificável, o XML é gerado mas **não é salvo**
""")

    with st.expander("Banco de Certificados — O que faz e quando usar"):
        st.markdown("""
- Visualiza métricas gerais do banco (certificados, itens, marcas)
- Permite enviar um PDF manualmente com confirmação antes de salvar
- Pesquisa item por referência/modelo ou código de barras
- Filtra todos os itens de uma marca específica
- Consulta histórico de alterações por IP-BRI (itens novos, removidos, campos alterados)
- Permite **remover** um certificado e seus itens permanentemente
- Exporta o banco completo em CSV
""")

    with st.expander("Registros — O que faz e quando usar"):
        st.markdown("""
- Importa o Excel de registros de um cliente (fábricas, CE-BRI, famílias, números de registro)
- Cada aba do Excel é tratada como uma fábrica — você pode renomear antes de salvar
- Os registros são necessários para que o Preenchimento de Confirmação consiga encontrar o número de registro e endereço da fábrica
- Permite editar o endereço de uma fábrica já cadastrada sem precisar reimportar tudo
""")

    with st.expander("Preenchimento de Confirmação — O que faz e quando usar"):
        st.markdown("""
- Recebe um Excel com células coloridas:
  - 🟢 **Verde** = referência do produto (MODELO). O sistema busca no banco por este valor
  - 🟡 **Amarelo** = campo a ser preenchido (MARCA, NOME, CODIGO, REGISTRO, ENDERECO, etc.)
  - 🔵 **Azul** = cabeçalho da coluna (diz qual campo preencher nas amarelas abaixo)
- A busca é feita primeiro no **Sistema 5**, depois nos **Certificados**
- **Pré-requisito:** o banco precisa ter certificados e registros cadastrados
""")

    with st.expander("Pesquisa Avançada — O que faz e quando usar"):
        st.markdown("""
- Pesquisa bidirecional: dado um IP-BRI, retorna o registro/fábrica; dado um registro, retorna o IP-BRI
- Útil para conferir rapidamente se um certificado está vinculado a um cliente
""")

    with st.expander("Sistema 5 — O que faz e quando usar"):
        st.markdown("""
- Módulo para processos mais recentes que ainda não têm certificado IP-BRI oficial
- Importa Excel de inclusões/manutenções/recertificações com IP no nome do arquivo
- Os itens salvos aqui também são buscados pelo Preenchimento de Confirmação
- Organizado por categoria → cliente → fábrica → processo
""")

    st.divider()

    st.subheader("💾 Backup")
    st.markdown("""
O banco de dados fica salvo no servidor do Streamlit. **Se o app reiniciar ou ficar inativo por muito tempo, os dados podem ser perdidos.**

**Boas práticas:**
- Baixe o backup geral pelo menu lateral sempre que terminar uma sessão de cadastro
- Guarde o arquivo `.db` em um local seguro
- Para restaurar, use o menu lateral → "Restaurar backup" e envie o arquivo `.db`
""")
    st.warning("⚠️ O Streamlit Cloud pode reiniciar o app a qualquer momento. Faça backup com frequência.")

with aba1:
    st.title("PDF → XML")

    uploaded_file = st.file_uploader(
        "Envie o PDF",
        type="pdf",
        key="xml_pdf"
    )

    if uploaded_file:
        with tempfile.TemporaryDirectory() as tmpdir:
            pdf_path = Path(tmpdir) / uploaded_file.name
            pdf_path.write_bytes(uploaded_file.read())

            rows = parse_pdf(pdf_path)

            if not rows:
                st.error("Nenhum item encontrado")
                st.stop()

            dados_certificado = extrair_dados_certificado(pdf_path)

            df = pd.DataFrame(
                rows,
                columns=[
                    "ORDEM",
                    "MARCA",
                    "MODELO",
                    "NOME",
                    "CODIGO"
                ]
            )

            st.success("Itens extraídos")

            st.dataframe(
                df,
                use_container_width=True
            )

            st.info(
                f"""
IP-BRI: {dados_certificado['ip_bri']}

CE-BRI: {dados_certificado['ce_bri']}

FAMÍLIA: {dados_certificado.get('familia')}

REV: {dados_certificado['rev']}
"""
            )

            registro_vinculado = buscar_registro_por_certificado(
                dados_certificado.get("ce_bri"),
                dados_certificado.get("familia")
            )

            if registro_vinculado.empty:
                st.warning("Nenhum registro vinculado encontrado para este CE-BRI + FAMÍLIA.")
            else:
                st.success("Registro vinculado encontrado ✅")
                st.dataframe(registro_vinculado, use_container_width=True)

            duplicados = verificar_codigos_duplicados(df)

            if not duplicados.empty:
                st.warning("Foram encontrados códigos duplicados neste PDF.")
                st.dataframe(duplicados, use_container_width=True)

            if not dados_certificado["ip_bri"]:
                st.warning("⚠️ IP-BRI não encontrado neste PDF. O XML foi gerado, mas o certificado **não foi salvo no banco**.")

            status, mensagem = salvar_ou_atualizar_certificado(
                dados_certificado,
                rows,
                uploaded_file.name
            )

            if status == "novo":
                st.success(mensagem)
            elif status == "atualizado":
                st.warning(mensagem)
            elif status == "ignorado":
                st.info(mensagem)
            else:
                st.error(mensagem)

            xml_comma = gerar_xml(df, ",")
            xml_dot = gerar_xml(df, ".")

            zip_path = Path(tmpdir) / "resultado_xml.zip"

            with zipfile.ZipFile(zip_path, "w") as zipf:
                zipf.writestr(
                    "xml_virgula.xml",
                    xml_comma.encode(
                        "ISO-8859-1",
                        errors="replace"
                    )
                )

                zipf.writestr(
                    "xml_ponto.xml",
                    xml_dot.encode(
                        "ISO-8859-1",
                        errors="replace"
                    )
                )

            with open(zip_path, "rb") as f:
                st.download_button(
                    "Baixar XMLs",
                    f,
                    "resultado_xml.zip"
                )

# ==========================================
# ABA 2 - BANCO
# ==========================================

with aba2:
    st.title("Banco de Certificados")

    # --- Dashboard de métricas ---
    try:
        total_certs = cursor.execute("SELECT COUNT(*) FROM certificados").fetchone()[0]
        total_itens_banco = cursor.execute("SELECT COUNT(*) FROM itens").fetchone()[0]
        total_marcas = cursor.execute("SELECT COUNT(DISTINCT marca) FROM itens WHERE marca IS NOT NULL AND marca != ''").fetchone()[0]
        ultimo_cert = cursor.execute("SELECT ip_bri FROM certificados ORDER BY id DESC LIMIT 1").fetchone()
        ultimo_cert_str = ultimo_cert[0] if ultimo_cert else "—"

        col_m1, col_m2, col_m3, col_m4 = st.columns(4)
        col_m1.metric("📄 Certificados", total_certs)
        col_m2.metric("📦 Itens", total_itens_banco)
        col_m3.metric("🏷️ Marcas distintas", total_marcas)
        col_m4.metric("🆕 Último cadastro", ultimo_cert_str)
        st.divider()
    except Exception:
        pass

    st.subheader("Backup do Banco")

    backup_upload = st.file_uploader(
        "Importar backup certificados.db",
        type=["db"],
        key="backup_db"
    )

    if backup_upload:
        st.warning("⚠️ Isso irá **substituir completamente** o banco atual. Esta ação não pode ser desfeita.")
        confirmar_backup_aba2 = st.checkbox("Confirmo que quero substituir o banco atual", key="confirmar_backup_aba2")
        if confirmar_backup_aba2:
            conn.close()
            with open(DB_PATH, "wb") as f:
                f.write(backup_upload.read())
            st.success("Backup importado com sucesso. Recarregue o app.")
            st.stop()

    if Path(DB_PATH).exists():
        with open(DB_PATH, "rb") as f:
            st.download_button(
                "Baixar backup atualizado",
                f,
                "certificados.db",
                "application/octet-stream"
            )

    st.divider()

    st.subheader("Exportar banco completo")

    todos = exportar_todos_itens()

    if todos.empty:
        st.info("Nenhum item cadastrado ainda.")
    else:
        st.download_button(
            "Baixar banco completo",
            todos.to_csv(
                index=False,
                sep=";"
            ).encode(
                "ISO-8859-1",
                errors="replace"
            ),
            "banco_completo.csv",
            "text/csv"
        )

    st.divider()

    st.subheader("Enviar certificado para o banco")

    banco_file = st.file_uploader(
        "Envie um certificado PDF",
        type="pdf",
        key="banco_pdf"
    )

    if banco_file:
        with tempfile.TemporaryDirectory() as tmpdir:
            pdf_path = Path(tmpdir) / banco_file.name
            pdf_path.write_bytes(banco_file.read())

            dados_certificado = extrair_dados_certificado(pdf_path)
            rows = parse_pdf(pdf_path)

            col1, col2, col3, col4, col5, col6 = st.columns(6)

            col1.metric("IP-BRI", dados_certificado["ip_bri"] or "Não encontrado")
            col2.metric("CE-BRI", dados_certificado["ce_bri"] or "Não encontrado")
            col3.metric("FAM", dados_certificado.get("familia") or "Não encontrada")
            col4.metric("REV", dados_certificado["rev"])
            col5.metric("Produto", dados_certificado["produto"] or "Não encontrado")
            col6.metric("Emissão", dados_certificado["data_emissao"] or "Não encontrado")

            registro_vinculado = buscar_registro_por_certificado(
                dados_certificado.get("ce_bri"),
                dados_certificado.get("familia")
            )

            if registro_vinculado.empty:
                st.warning("Nenhum registro vinculado encontrado para este CE-BRI + FAMÍLIA.")
            else:
                st.success("Registro vinculado encontrado ✅")
                st.dataframe(registro_vinculado, use_container_width=True)

            if rows:
                df_preview = pd.DataFrame(
                    rows,
                    columns=[
                        "ORDEM",
                        "MARCA",
                        "MODELO",
                        "NOME",
                        "CODIGO"
                    ]
                )

                st.dataframe(df_preview, use_container_width=True)

                duplicados = verificar_codigos_duplicados(df_preview)

                if not duplicados.empty:
                    st.warning("Foram encontrados códigos duplicados.")
                    st.dataframe(duplicados, use_container_width=True)

                if st.button("Salvar / Atualizar banco"):
                    status, mensagem = salvar_ou_atualizar_certificado(
                        dados_certificado,
                        rows,
                        banco_file.name
                    )

                    if status in ["novo", "atualizado"]:
                        st.success(mensagem)
                    elif status == "ignorado":
                        st.info(mensagem)
                    else:
                        st.error(mensagem)

    st.divider()

    st.subheader("Pesquisar item no banco")

    tipo_busca = st.selectbox(
        "Pesquisar por",
        ["Referência / Modelo", "Código de Barras"],
        key="tipo_busca"
    )

    termo_busca = st.text_input(
        "Digite a referência/modelo ou código",
        key="busca_item"
    )

    if termo_busca:
        campo = "i.modelo" if tipo_busca == "Referência / Modelo" else "i.codigo"

        resultado_item = pd.read_sql_query(f"""
        SELECT
            i.marca AS marca,
            i.modelo AS modelo,
            i.nome AS nome,
            i.codigo AS codigo,
            c.ip_bri AS ip_bri,
            c.ce_bri AS ce_bri,
            r.registro AS registro,
            c.familia AS fam,
            r.fabrica AS fabrica,
            c.rev AS rev,
            c.produto AS produto,
            c.data_emissao AS data_emissao,
            i.ordem AS ordem
        FROM itens i
        INNER JOIN certificados c
        ON c.id = i.certificado_id
        LEFT JOIN registros r
        ON UPPER(r.ce_bri) = UPPER(c.ce_bri)
        AND r.familia = c.familia
        WHERE {campo} LIKE ?
        ORDER BY i.marca, i.modelo, c.ip_bri, i.ordem
        """, conn, params=(f"%{termo_busca}%",))

        if resultado_item.empty:
            st.warning("Nenhum item encontrado.")
        else:
            st.success("Item encontrado ✅")
            st.dataframe(resultado_item, use_container_width=True)

            st.download_button(
                "Baixar resultado da pesquisa em CSV",
                resultado_item.to_csv(index=False, sep=";").encode(
                    "ISO-8859-1",
                    errors="replace"
                ),
                "resultado_pesquisa_item.csv",
                "text/csv"
            )

    st.divider()

    st.subheader("Itens Repetidos")

    if st.button("Ver itens repetidos"):
        itens_repetidos = pd.read_sql_query("""
        SELECT
            i.marca AS marca,
            i.modelo AS modelo,
            i.nome AS nome,
            i.codigo AS codigo,
            c.ip_bri AS ip_bri,
            c.ce_bri AS ce_bri,
            r.registro AS registro,
            c.familia AS fam,
            r.fabrica AS fabrica,
            c.rev AS rev,
            c.produto AS produto,
            c.data_emissao AS data_emissao,
            i.ordem AS ordem
        FROM itens i
        INNER JOIN certificados c
        ON c.id = i.certificado_id
        LEFT JOIN registros r
        ON UPPER(r.ce_bri) = UPPER(c.ce_bri)
        AND r.familia = c.familia
        WHERE i.codigo IN (
            SELECT codigo
            FROM itens
            WHERE codigo IS NOT NULL
            AND codigo != ''
            GROUP BY codigo
            HAVING COUNT(*) > 1
        )
        ORDER BY i.codigo, i.marca, i.modelo
        """, conn)

        if itens_repetidos.empty:
            st.success("Nenhum item repetido encontrado ✅")
        else:
            st.warning("Itens repetidos encontrados.")
            st.dataframe(itens_repetidos, use_container_width=True)

            st.download_button(
                "Baixar itens repetidos em CSV",
                itens_repetidos.to_csv(index=False, sep=";").encode(
                    "ISO-8859-1",
                    errors="replace"
                ),
                "itens_repetidos.csv",
                "text/csv"
            )

    st.divider()

    st.subheader("Filtro geral por Marca")

    marcas_banco = pd.read_sql_query("""
    SELECT DISTINCT marca
    FROM itens
    WHERE marca IS NOT NULL
    AND marca != ''
    ORDER BY marca
    """, conn)

    if marcas_banco.empty:
        st.info("Nenhuma marca cadastrada.")
    else:
        marca_filtro = st.selectbox(
            "Selecione uma marca",
            marcas_banco["marca"].tolist(),
            key="marca_filtro"
        )

        resultado_marca = pd.read_sql_query("""
        SELECT
            i.marca AS marca,
            i.modelo AS modelo,
            i.nome AS nome,
            i.codigo AS codigo,
            c.ip_bri AS ip_bri,
            c.ce_bri AS ce_bri,
            r.registro AS registro,
            c.familia AS fam,
            r.fabrica AS fabrica,
            c.rev AS rev,
            c.produto AS produto,
            c.data_emissao AS data_emissao,
            i.ordem AS ordem
        FROM itens i
        INNER JOIN certificados c
        ON c.id = i.certificado_id
        LEFT JOIN registros r
        ON UPPER(r.ce_bri) = UPPER(c.ce_bri)
        AND r.familia = c.familia
        WHERE i.marca = ?
        ORDER BY i.modelo, c.ip_bri
        """, conn, params=(marca_filtro,))

        st.dataframe(resultado_marca, use_container_width=True)

        st.download_button(
            "Baixar CSV da marca",
            resultado_marca.to_csv(index=False, sep=";").encode(
                "ISO-8859-1",
                errors="replace"
            ),
            f"{marca_filtro}.csv",
            "text/csv"
        )

    st.divider()

    st.subheader("Histórico por IP-BRI")

    ip_historico = st.text_input(
        "Digite o IP-BRI",
        key="historico_ip"
    )

    if ip_historico:
        historico = pd.read_sql_query("""
        SELECT
            ip_bri,
            rev_antiga,
            rev_nova,
            modelo,
            codigo,
            tipo_alteracao,
            campo_alterado,
            valor_antigo,
            valor_novo,
            arquivo_pdf,
            data_hora
        FROM historico_alteracoes
        WHERE ip_bri LIKE ?
        ORDER BY id DESC
        """, conn, params=(f"%{ip_historico}%",))

        if historico.empty:
            st.warning("Nenhum histórico encontrado.")
        else:
            grupos = historico.groupby(
                [
                    "rev_antiga",
                    "rev_nova",
                    "arquivo_pdf",
                    "data_hora"
                ],
                dropna=False
            )

            for (rev_antiga, rev_nova, arquivo_pdf, data_hora), grupo in grupos:
                with st.expander(
                    f"REV {rev_antiga} → REV {rev_nova} | {data_hora} | {arquivo_pdf}"
                ):
                    col1, col2, col3 = st.columns(3)

                    col1.metric(
                        "Itens novos",
                        len(grupo[grupo["tipo_alteracao"] == "ITEM_NOVO"])
                    )

                    col2.metric(
                        "Itens removidos",
                        len(grupo[grupo["tipo_alteracao"] == "ITEM_REMOVIDO"])
                    )

                    col3.metric(
                        "Campos alterados",
                        len(grupo[grupo["tipo_alteracao"] == "CAMPO_ALTERADO"])
                    )

                    for _, row in grupo.iterrows():
                        tipo = row["tipo_alteracao"]

                        if tipo == "CERTIFICADO_NOVO":
                            st.info(f"📄 Certificado novo: {row['valor_novo']}")

                        elif tipo == "CERTIFICADO_ATUALIZADO":
                            st.success(f"🔄 Atualizado: {row['valor_antigo']} → {row['valor_novo']}")

                        elif tipo == "REV_IGNORADA":
                            st.warning(f"⚠️ Revisão ignorada: {row['valor_antigo']} | {row['valor_novo']}")

                        elif tipo == "ITEM_REMOVIDO":
                            st.error(
                                f"🗑️ ITEM REMOVIDO\n\n"
                                f"Modelo: {row['modelo']}\n\n"
                                f"Código: {row['codigo']}\n\n"
                                f"Antes: {row['valor_antigo']}"
                            )

                        elif tipo == "ITEM_NOVO":
                            st.success(
                                f"➕ ITEM NOVO\n\n"
                                f"Modelo: {row['modelo']}\n\n"
                                f"Código: {row['codigo']}\n\n"
                                f"Novo: {row['valor_novo']}"
                            )

                        elif tipo == "CAMPO_ALTERADO":
                            st.warning(
                                f"✏️ CAMPO ALTERADO\n\n"
                                f"Modelo: {row['modelo']}\n\n"
                                f"Código: {row['codigo']}\n\n"
                                f"Campo: {row['campo_alterado']}\n\n"
                                f"Antes: {row['valor_antigo']}\n\n"
                                f"Depois: {row['valor_novo']}"
                            )

            st.download_button(
                "Baixar histórico CSV",
                historico.to_csv(index=False, sep=";").encode(
                    "ISO-8859-1",
                    errors="replace"
                ),
                "historico.csv",
                "text/csv"
            )

    st.divider()

    st.subheader("🗑️ Remover Certificado do Banco")
    st.caption("Use com cuidado. Remove permanentemente o certificado e todos os seus itens do banco.")

    ip_deletar = st.text_input(
        "Digite o IP-BRI exato para remover",
        placeholder="Ex: IP-BRI-0533/2023-15",
        key="ip_deletar"
    )

    if ip_deletar:
        cert_para_deletar = cursor.execute(
            "SELECT id, ip_bri, rev, produto FROM certificados WHERE ip_bri = ?",
            (ip_deletar.strip(),)
        ).fetchone()

        if not cert_para_deletar:
            st.warning("Certificado não encontrado no banco.")
        else:
            qtd_itens_cert = cursor.execute(
                "SELECT COUNT(*) FROM itens WHERE certificado_id = ?",
                (cert_para_deletar[0],)
            ).fetchone()[0]

            st.error(
                f"**{cert_para_deletar[1]}** | REV {cert_para_deletar[2]} | "
                f"{cert_para_deletar[3] or '—'} | {qtd_itens_cert} itens vinculados"
            )

            confirmar_delete = st.checkbox(
                f"Confirmo que quero remover permanentemente este certificado e seus {qtd_itens_cert} itens",
                key="confirmar_delete_cert"
            )

            if confirmar_delete:
                if st.button("🗑️ Remover do banco", key="btn_deletar_cert"):
                    try:
                        cursor.execute("DELETE FROM itens WHERE certificado_id = ?", (cert_para_deletar[0],))
                        cursor.execute("DELETE FROM certificados WHERE id = ?", (cert_para_deletar[0],))
                        conn.commit()
                        st.success(f"Certificado {cert_para_deletar[1]} e seus itens foram removidos com sucesso ✅")
                    except Exception as e:
                        st.error(f"Erro ao remover: {e}")

# ==========================================
# ABA 3 - REGISTROS
# ==========================================

with aba3:
    st.title("Registros")

    st.subheader("Enviar Excel de Registros")

    cliente_base = st.text_input(
        "Nome da Base / Cliente",
        value="BOLSA",
        key="cliente_base_registros"
    )

    with st.expander("ℹ️ Como funciona esta aba"):
        st.markdown("""
- Envie o Excel de registros do cliente
- Cada **aba do Excel** é tratada como uma fábrica separada
- Você pode renomear cada fábrica e informar o endereço antes de salvar
- Os registros vinculam o **CE-BRI** ao número de **registro** e à **fábrica**, permitindo que o Preenchimento de Confirmação preencha esses campos automaticamente
""")

    registro_excel = st.file_uploader(
        "Envie o Excel de Registros",
        type=["xlsx", "xls"],
        key="excel_registros"
    )

    if registro_excel:
        try:
            excel = pd.ExcelFile(registro_excel)
            abas = excel.sheet_names

            st.success(f"{len(abas)} abas encontradas ✅")

            st.subheader("Nomear Fábricas")
            st.caption("Se quiser, altere o nome de cada fábrica antes de salvar. Exemplo: F01 - MOHNISH / CE-BRI-XXXX")

            nomes_fabricas = {}
            enderecos_fabricas_por_aba = {}

            for aba in abas:
                nome_sugerido = str(aba)

                nomes_fabricas[aba] = st.text_input(
                    f"Nome da fábrica para a aba: {aba}",
                    value=nome_sugerido,
                    key=f"nome_fabrica_{cliente_base}_{aba}"
                )

                enderecos_fabricas_por_aba[aba] = st.text_area(
                    f"Endereço da fábrica para a aba: {aba}",
                    value="",
                    key=f"endereco_fabrica_{cliente_base}_{aba}",
                    height=80
                )

            todas_fabricas = []

            for aba in abas:
                df_filtrado = ler_registros_aba(
                    registro_excel,
                    aba
                )

                if df_filtrado is None or df_filtrado.empty:
                    continue

                nome_final_fabrica = clean(nomes_fabricas.get(aba, aba)) or str(aba)
                endereco_final_fabrica = clean(enderecos_fabricas_por_aba.get(aba, ""))

                # Mantém o CE-BRI extraído da aba original, mas salva a fábrica com o nome escolhido.
                df_filtrado["FABRICA"] = nome_final_fabrica
                df_filtrado["ENDERECO_FABRICA"] = endereco_final_fabrica

                st.subheader(f"Prévia: {cliente_base} → {nome_final_fabrica}")

                st.dataframe(
                    df_filtrado,
                    use_container_width=True
                )

                todas_fabricas.append(df_filtrado)

            if todas_fabricas:
                banco_registros = pd.concat(
                    todas_fabricas,
                    ignore_index=True
                )

                st.divider()

                st.subheader("Banco Geral de Registros que será salvo")

                st.dataframe(
                    banco_registros,
                    use_container_width=True
                )

                if st.button("Salvar registros desta base", key="salvar_registros_base"):
                    enderecos_para_salvar = {}
                    if "ENDERECO_FABRICA" in banco_registros.columns:
                        for _, linha_endereco in banco_registros.drop_duplicates(subset=["FABRICA"]).iterrows():
                            enderecos_para_salvar[str(linha_endereco.get("FABRICA", ""))] = clean(linha_endereco.get("ENDERECO_FABRICA", ""))

                    total_salvo = salvar_registros_no_banco(
                        banco_registros,
                        registro_excel.name,
                        cliente_base,
                        enderecos_para_salvar
                    )

                    st.success(f"{total_salvo} registros salvos no banco para a base {cliente_base} ✅")

                st.download_button(
                    "Baixar registros tratados CSV",
                    banco_registros.to_csv(
                        index=False,
                        sep=";"
                    ).encode(
                        "ISO-8859-1",
                        errors="replace"
                    ),
                    "registros_tratados.csv",
                    "text/csv"
                )

            else:
                st.warning("Nenhum registro válido encontrado no Excel.")

        except Exception as e:
            st.error(f"Erro ao ler Excel: {e}")

    st.divider()

    st.subheader("Backup dos Registros / Banco")

    backup_registro_upload = st.file_uploader(
        "Importar backup certificados.db",
        type=["db"],
        key="backup_db_registros"
    )

    if backup_registro_upload:
        st.warning("⚠️ Isso irá **substituir completamente** o banco atual. Esta ação não pode ser desfeita.")
        confirmar_backup_aba3 = st.checkbox("Confirmo que quero substituir o banco atual", key="confirmar_backup_aba3")
        if confirmar_backup_aba3:
            conn.close()
            with open(DB_PATH, "wb") as f:
                f.write(backup_registro_upload.read())
            st.success("Backup importado com sucesso. Recarregue o app.")
            st.stop()

    if Path(DB_PATH).exists():
        with open(DB_PATH, "rb") as f:
            st.download_button(
                "Baixar backup atualizado",
                f,
                "certificados.db",
                "application/octet-stream",
                key="download_backup_registros"
            )

    registros_salvos = exportar_registros_banco()

    if not registros_salvos.empty:
        st.download_button(
            "Baixar registros salvos CSV",
            registros_salvos.to_csv(index=False, sep=";").encode(
                "ISO-8859-1",
                errors="replace"
            ),
            "registros_salvos.csv",
            "text/csv",
            key="download_registros_salvos"
        )

    st.divider()

    st.subheader("Visualizar Bases")

    bases_existentes = pd.read_sql_query("""
    SELECT DISTINCT cliente_base
    FROM registros
    WHERE cliente_base IS NOT NULL
    AND cliente_base != ''
    ORDER BY cliente_base
    """, conn)

    if bases_existentes.empty:
        st.info("Nenhuma base salva ainda.")
    else:
        base_selecionada = st.selectbox(
            "Selecione a base",
            bases_existentes["cliente_base"].tolist(),
            key="visualizar_base"
        )

        registros_base = pd.read_sql_query("""
        SELECT
            cliente_base,
            fabrica,
            ce_bri,
            familia,
            registro,
            endereco_fabrica
        FROM registros
        WHERE cliente_base = ?
        ORDER BY fabrica, CAST(familia AS INTEGER)
        """, conn, params=(base_selecionada,))

        fabricas = registros_base["fabrica"].unique().tolist()

        for fabrica in fabricas:
            bloco = registros_base[
                registros_base["fabrica"] == fabrica
            ]

            with st.expander(f"{base_selecionada} → {fabrica}", expanded=False):
                st.dataframe(
                    bloco[[
                        "familia",
                        "registro",
                        "ce_bri",
                        "endereco_fabrica"
                    ]],
                    use_container_width=True
                )

                st.download_button(
                    f"Baixar CSV - {fabrica}",
                    bloco.to_csv(index=False, sep=";").encode(
                        "ISO-8859-1",
                        errors="replace"
                    ),
                    f"{base_selecionada}_{str(fabrica).replace('/', '-')}.csv",
                    "text/csv",
                    key=f"download_{base_selecionada}_{fabrica}"
                )


    st.divider()

    st.subheader("Editar endereço de fábrica já salva")

    bases_para_editar = pd.read_sql_query("""
    SELECT DISTINCT cliente_base
    FROM registros
    WHERE cliente_base IS NOT NULL
    AND cliente_base != ''
    ORDER BY cliente_base
    """, conn)

    if bases_para_editar.empty:
        st.info("Nenhuma base disponível para edição de endereço.")
    else:
        base_editar = st.selectbox(
            "Base / Cliente para editar endereço",
            bases_para_editar["cliente_base"].tolist(),
            key="editar_endereco_base"
        )

        fabricas_para_editar = pd.read_sql_query("""
        SELECT
            fabrica,
            MAX(ce_bri) AS ce_bri,
            MAX(endereco_fabrica) AS endereco_fabrica
        FROM registros
        WHERE cliente_base = ?
        GROUP BY fabrica
        ORDER BY fabrica
        """, conn, params=(base_editar,))

        if fabricas_para_editar.empty:
            st.info("Nenhuma fábrica encontrada nessa base.")
        else:
            opcoes_fabricas = fabricas_para_editar["fabrica"].tolist()

            fabrica_editar = st.selectbox(
                "Fábrica para editar endereço",
                opcoes_fabricas,
                key="editar_endereco_fabrica"
            )

            linha_fabrica = fabricas_para_editar[
                fabricas_para_editar["fabrica"] == fabrica_editar
            ].iloc[0]

            st.caption(f"CE-BRI vinculado: {linha_fabrica.get('ce_bri', '')}")

            endereco_atual = linha_fabrica.get("endereco_fabrica", "")

            if pd.isna(endereco_atual):
                endereco_atual = ""

            novo_endereco = st.text_area(
                "Endereço da fábrica",
                value=str(endereco_atual),
                height=120,
                key="editar_endereco_texto"
            )

            if st.button("Salvar endereço desta fábrica", key="salvar_endereco_fabrica_existente"):
                total_alterado = atualizar_endereco_fabrica_existente(
                    base_editar,
                    fabrica_editar,
                    novo_endereco
                )

                st.success(f"Endereço atualizado em {total_alterado} registros da fábrica {fabrica_editar} ✅")


# ==========================================
# ABA 4 - PREENCHIMENTO DE CONFIRMAÇÃO
# ==========================================

with aba4:
    st.title("Preenchimento de Confirmação")

    with st.expander("ℹ️ Como funciona esta aba"):
        st.markdown("""
- Envie o Excel de confirmação com células coloridas:
  - 🟢 **Verde** → referência do produto (MODELO). O sistema busca este valor no banco
  - 🟡 **Amarelo** → campo a ser preenchido (MARCA, NOME, CODIGO, REGISTRO, ENDERECO...)
  - 🔵 **Azul** → cabeçalho da coluna (define o que preencher nas amarelas abaixo)
- Não é necessário ter uma coluna chamada "REF" — qualquer célula verde é considerada referência
- A busca é feita primeiro no **Sistema 5**, depois nos **Certificados**
- **Pré-requisito:** o banco precisa ter certificados e registros cadastrados para o preenchimento funcionar
""")

    # Aviso de banco vazio — aba 4 não funciona sem dados
    try:
        _total_cert_aba4 = cursor.execute("SELECT COUNT(*) FROM certificados").fetchone()[0]
        _total_s5_aba4   = cursor.execute("SELECT COUNT(*) FROM sistema5_itens").fetchone()[0]
        _total_reg_aba4  = cursor.execute("SELECT COUNT(*) FROM registros").fetchone()[0]

        if _total_cert_aba4 == 0 and _total_s5_aba4 == 0:
            st.error(
                "⛔ O banco não tem nenhum certificado cadastrado. "
                "O preenchimento não vai encontrar nenhuma referência.\n\n"
                "**O que fazer:** vá à aba **PDF → XML** e envie os certificados primeiro."
            )
        elif _total_reg_aba4 == 0:
            st.warning(
                "⚠️ O banco tem certificados, mas **nenhum registro de fábrica** cadastrado. "
                "Os campos REGISTRO e ENDERECO não serão preenchidos.\n\n"
                "**O que fazer:** vá à aba **Registros** e importe o Excel de registros do cliente."
            )
        else:
            st.success(
                f"Banco pronto: {_total_cert_aba4} certificado(s) + {_total_s5_aba4} item(ns) do Sistema 5 + {_total_reg_aba4} registro(s) de fábrica ✅"
            )
    except Exception:
        pass

    arquivo_confirmacao = st.file_uploader(
        "Envie o Excel de confirmação",
        type=["xlsx", "xlsm"],
        key="excel_confirmacao"
    )

    if arquivo_confirmacao:
        try:
            saida_excel, total_preenchidos, nao_encontrados = preencher_excel_confirmacao(
                arquivo_confirmacao
            )

            st.success(f"Preenchimento concluído ✅ {total_preenchidos} células preenchidas.")

            if nao_encontrados:
                st.warning("Algumas referências verdes não foram encontradas no banco.")
                st.dataframe(
                    pd.DataFrame({"referencia_nao_encontrada": sorted(set(nao_encontrados))}),
                    use_container_width=True
                )

            st.download_button(
                "Baixar Excel preenchido",
                saida_excel,
                "confirmacao_preenchida.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Erro ao preencher Excel: {e}")


# ==========================================
# ABA 5 - PESQUISA AVANÇADA
# ==========================================

with aba5:
    st.title("Pesquisa Avançada")

    with st.expander("ℹ️ Como funciona esta aba"):
        st.markdown("""
- Pesquise em **duas direções**:
  - Dado um **IP-BRI** → descubra o registro, fábrica e endereço vinculados
  - Dado um **REGISTRO** → descubra o IP-BRI e o fabricante
- Útil para conferir rapidamente se um certificado está vinculado a um cliente antes do preenchimento
""")

    tipo_pesquisa_avancada = st.radio(
        "O que você quer pesquisar?",
        [
            "Quero saber o REGISTRO de um IP-BRI",
            "Quero saber o IP-BRI de um REGISTRO"
        ],
        key="tipo_pesquisa_avancada"
    )

    if tipo_pesquisa_avancada == "Quero saber o REGISTRO de um IP-BRI":
        ip_pesquisa = st.text_input(
            "Digite o IP-BRI",
            placeholder="Ex: IP-BRI-0533/2023-15 ou 0533/2023-15",
            key="pesquisa_avancada_ip"
        )

        if ip_pesquisa:
            resultado_ip = pesquisar_por_ip_bri(ip_pesquisa)

            if resultado_ip.empty:
                st.warning("Nenhum registro encontrado para este IP-BRI.")
            else:
                st.success(f"{len(resultado_ip)} resultado(s) encontrado(s) ✅")
                st.dataframe(resultado_ip, use_container_width=True)

                st.download_button(
                    "Baixar resultado em CSV",
                    resultado_ip.to_csv(index=False, sep=";").encode(
                        "ISO-8859-1",
                        errors="replace"
                    ),
                    "pesquisa_por_ip_bri.csv",
                    "text/csv",
                    key="download_pesquisa_ip_bri"
                )

    else:
        registro_pesquisa = st.text_input(
            "Digite o REGISTRO",
            placeholder="Ex: 003889/2023",
            key="pesquisa_avancada_registro"
        )

        if registro_pesquisa:
            resultado_registro = pesquisar_por_registro(registro_pesquisa)

            if resultado_registro.empty:
                st.warning("Nenhum IP-BRI encontrado para este registro.")
            else:
                st.success(f"{len(resultado_registro)} resultado(s) encontrado(s) ✅")
                st.dataframe(resultado_registro, use_container_width=True)

                st.download_button(
                    "Baixar resultado em CSV",
                    resultado_registro.to_csv(index=False, sep=";").encode(
                        "ISO-8859-1",
                        errors="replace"
                    ),
                    "pesquisa_por_registro.csv",
                    "text/csv",
                    key="download_pesquisa_registro"
                )

# ==========================================
# ABA 6 - SISTEMA 5
# ==========================================

with aba6:
    st.title("Sistema 5")

    if st.button("Limpar itens DESCONSIDERADOS já cadastrados", key="limpar_desconsiderados_s5"):
        removidos_desconsiderados = limpar_itens_desconsiderados_sistema5()
        st.success(f"{removidos_desconsiderados} item(ns) desconsiderado(s) removido(s) do Sistema 5 ✅")

    st.divider()
    st.subheader("🔎 Pesquisa Global Sistema 5")

    termo_pesquisa_s5 = st.text_input(
        "Pesquisar item no Sistema 5",
        placeholder="Digite modelo, código, marca, descrição, IP, CE-BRI, fábrica ou cliente",
        key="pesquisa_global_sistema5"
    )

    if st.button("Pesquisar no Sistema 5", key="botao_pesquisa_global_s5"):
        if not clean(termo_pesquisa_s5):
            st.warning("Digite algo para pesquisar.")
        else:
            resultado_global_s5 = pesquisar_global_sistema5(termo_pesquisa_s5)

            if resultado_global_s5.empty:
                st.warning("Nenhum item encontrado no Sistema 5.")
            else:
                st.success(f"{len(resultado_global_s5)} item(ns) encontrado(s) no Sistema 5 ✅")
                st.dataframe(resultado_global_s5, use_container_width=True)

                st.download_button(
                    "Baixar resultado Sistema 5 CSV",
                    resultado_global_s5.to_csv(index=False, sep=";").encode(
                        "ISO-8859-1",
                        errors="replace"
                    ),
                    "pesquisa_global_sistema5.csv",
                    "text/csv",
                    key="download_pesquisa_global_s5"
                )


    with st.expander("ℹ️ Como funciona esta aba"):
        st.markdown("""
- Módulo para processos que ainda não têm certificado IP-BRI oficial emitido
- Importa Excel de **inclusões, manutenções ou recertificações** — o tipo é detectado pelo nome do arquivo
- **O arquivo precisa ter IP no nome** (ex: `INCLUSÃO 06-01-26 F01 IP-0094-26.xlsx`)
- Cada aba do Excel é tratada como uma **família** pelo número presente no nome da aba
- Os itens salvos aqui também são consultados pelo **Preenchimento de Confirmação**
- Organizado por: categoria → cliente → fábrica → processo
""")

    categoria_s5 = st.radio(
        "Categoria",
        ["SISTEMA 5 NOVO PROJETO", "SISTEMA 5 PROPRIOS"],
        horizontal=True,
        key="s5_categoria"
    )

    cliente_s5 = st.text_input(
        "Cliente / Solicitante",
        value="BOLSA" if categoria_s5 == "SISTEMA 5 NOVO PROJETO" else "",
        key="s5_cliente"
    )

    st.subheader("Fábrica")

    fabricas_existentes_s5 = listar_fabricas_existentes_para_sistema5(cliente_s5)

    usar_fabrica_existente = False

    if not fabricas_existentes_s5.empty:
        usar_fabrica_existente = st.checkbox(
            "Usar fábrica já cadastrada nos Registros",
            value=True,
            key="s5_usar_fabrica_existente"
        )

    dados_fabrica_sugeridos = None

    if usar_fabrica_existente and not fabricas_existentes_s5.empty:
        opcoes_fabrica_s5 = []

        for _, linha_fab in fabricas_existentes_s5.iterrows():
            fab = clean(linha_fab.get("fabrica", ""))
            ce = clean(linha_fab.get("ce_bri", ""))
            opcoes_fabrica_s5.append(f"{fab} | {ce}")

        escolha_fabrica_s5 = st.selectbox(
            "Selecione uma fábrica já cadastrada",
            opcoes_fabrica_s5,
            key="s5_fabrica_existente_select"
        )

        fabrica_escolhida = escolha_fabrica_s5.split("|")[0].strip()
        ce_escolhido = escolha_fabrica_s5.split("|")[1].strip() if "|" in escolha_fabrica_s5 else ""

        dados_fabrica_sugeridos = buscar_dados_fabrica_existente(
            cliente_s5,
            fabrica_escolhida,
            ce_escolhido
        )

        fabrica_padrao = dados_fabrica_sugeridos.get("fabrica", fabrica_escolhida) if dados_fabrica_sugeridos else fabrica_escolhida
        ce_padrao = dados_fabrica_sugeridos.get("ce_bri", ce_escolhido) if dados_fabrica_sugeridos else ce_escolhido
        endereco_padrao = dados_fabrica_sugeridos.get("endereco_fabrica", "") if dados_fabrica_sugeridos else ""

        st.success("Dados da fábrica puxados dos Registros ✅")

    else:
        fabrica_padrao = ""
        ce_padrao = ""
        endereco_padrao = ""

    col_f1, col_f2 = st.columns(2)

    with col_f1:
        fabrica_s5 = st.text_input(
            "Fábrica",
            value=fabrica_padrao,
            placeholder="Ex: F01, F02, FÁBRICA 01...",
            key="s5_fabrica"
        )

        ce_bri_s5 = st.text_input(
            "CE-BRI da fábrica",
            value=ce_padrao,
            placeholder="Ex: CE-BRI-INNAC-02484-01A",
            key="s5_ce_bri"
        )

    with col_f2:
        endereco_s5 = st.text_area(
            "Endereço da fábrica",
            value="" if pd.isna(endereco_padrao) else str(endereco_padrao),
            placeholder="Cole aqui o endereço completo da fábrica",
            height=120,
            key="s5_endereco"
        )

    if clean(ce_bri_s5) and not clean(endereco_s5):
        dados_por_ce = buscar_dados_fabrica_existente(
            cliente_s5,
            "",
            ce_bri_s5
        )

        if dados_por_ce and clean(dados_por_ce.get("endereco_fabrica", "")):
            st.info("Existe endereço salvo para este CE-BRI nos Registros. Você pode copiar/preencher acima.")
            st.code(dados_por_ce.get("endereco_fabrica", ""))

    st.divider()
    st.subheader("Enviar Excel(s) de processo")

    arquivos_s5 = st.file_uploader(
        "Envie um ou mais Excels com IP no nome",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        key="s5_excel_multi"
    )

    if arquivos_s5:
        st.info(
            f"{len(arquivos_s5)} arquivo(s) selecionado(s). "
            "Cada arquivo será tratado como um processo separado, com validação e botão de salvar individual."
        )

        for idx_arquivo, arquivo_s5 in enumerate(arquivos_s5):
            st.divider()

            with st.expander(f"📄 {arquivo_s5.name}", expanded=True):
                ip_extraido = extrair_ip_processo(arquivo_s5.name)
                data_extraida = extrair_data_processo(arquivo_s5.name)
                tipo_detectado = detectar_tipo_processo(arquivo_s5.name)
                fabrica_detectada_nome = extrair_codigo_fabrica_nome(arquivo_s5.name)
                dados_fabrica_arquivo = buscar_fabrica_por_codigo_s5(cliente_s5, fabrica_detectada_nome) if fabrica_detectada_nome else None

                if not ip_extraido:
                    st.error(
                        "Arquivo ignorado: o nome do arquivo precisa conter IP. "
                        "Exemplo: INCLUSÃO 06-01-26 F01 IP-0094-26.xlsx"
                    )
                    continue

                st.success(f"IP identificado: {ip_extraido}")
                st.info(
                    f"Tipo detectado: {tipo_detectado} | "
                    f"Data detectada: {data_extraida or 'não encontrada'} | "
                    f"Fábrica detectada no arquivo: {fabrica_detectada_nome or 'não encontrada'}"
                )

                fabrica_final_arquivo = fabrica_s5
                ce_bri_final_arquivo = ce_bri_s5
                endereco_final_arquivo = endereco_s5

                if dados_fabrica_arquivo:
                    st.success(
                        f"Fábrica vinculada automaticamente pela base {cliente_s5}: "
                        f"{dados_fabrica_arquivo.get('fabrica', '')} | {dados_fabrica_arquivo.get('ce_bri', '')}"
                    )

                    usar_dados_auto_s5 = st.checkbox(
                        "Usar CE-BRI e endereço encontrados automaticamente para este arquivo",
                        value=True,
                        key=f"s5_usar_dados_auto_arquivo_{idx_arquivo}_{arquivo_s5.name}"
                    )

                    if usar_dados_auto_s5:
                        fabrica_final_arquivo = dados_fabrica_arquivo.get("fabrica", fabrica_final_arquivo)
                        ce_bri_final_arquivo = dados_fabrica_arquivo.get("ce_bri", ce_bri_final_arquivo)
                        endereco_final_arquivo = dados_fabrica_arquivo.get("endereco_fabrica", endereco_final_arquivo)

                elif fabrica_detectada_nome:
                    st.warning(
                        f"O arquivo indica {fabrica_detectada_nome}, mas não encontrei essa fábrica nos Registros da base {cliente_s5}."
                    )

                col_proc_1, col_proc_2 = st.columns(2)

                tipos = ["INCLUSAO", "MANUTENCAO", "RECERTIFICACAO", "INICIAL", "OUTROS"]

                with col_proc_1:
                    tipo_s5 = st.selectbox(
                        "Tipo de processo",
                        tipos,
                        index=tipos.index(tipo_detectado) if tipo_detectado in tipos else 0,
                        key=f"s5_tipo_{idx_arquivo}_{arquivo_s5.name}"
                    )

                    data_s5 = st.text_input(
                        "Data do processo",
                        value=data_extraida or "",
                        placeholder="AAAA-MM-DD",
                        key=f"s5_data_{idx_arquivo}_{arquivo_s5.name}"
                    )

                with col_proc_2:
                    st.text_input(
                        "Fábrica que será usada neste arquivo",
                        value=fabrica_final_arquivo,
                        key=f"s5_fabrica_final_{idx_arquivo}_{arquivo_s5.name}",
                        disabled=True
                    )

                    st.text_input(
                        "CE-BRI que será usado neste arquivo",
                        value=ce_bri_final_arquivo,
                        key=f"s5_ce_bri_final_{idx_arquivo}_{arquivo_s5.name}",
                        disabled=True
                    )

                duplicados_ip = ip_processo_ja_existe_sistema5(
                    ip_extraido,
                    cliente_s5,
                    fabrica_final_arquivo
                )

                if not duplicados_ip.empty:
                    st.warning("Já existe processo salvo com este IP para esta base/fábrica.")
                    st.dataframe(duplicados_ip, use_container_width=True)

                    substituir_ip = st.checkbox(
                        "Substituir processo existente com este IP antes de salvar",
                        value=False,
                        key=f"s5_substituir_ip_{idx_arquivo}_{arquivo_s5.name}"
                    )
                else:
                    substituir_ip = False

                try:
                    df_s5 = ler_excel_inclusao_sistema5(arquivo_s5)

                    if df_s5.empty:
                        st.warning("Nenhum item encontrado nesse Excel. Verifique se há coluna de MODELO / referência.")
                        continue

                    st.subheader("Prévia dos itens encontrados")
                    st.dataframe(df_s5, use_container_width=True)

                    if st.button(
                        f"Salvar este processo: {arquivo_s5.name}",
                        key=f"s5_salvar_processo_{idx_arquivo}_{arquivo_s5.name}"
                    ):
                        if not clean(cliente_s5) or not clean(fabrica_final_arquivo):
                            st.error("Informe cliente e fábrica antes de salvar.")
                        else:
                            if substituir_ip and not duplicados_ip.empty:
                                for _, proc_dup in duplicados_ip.iterrows():
                                    deletar_processo_sistema5(proc_dup["arquivo_id"])

                            total = salvar_inclusao_sistema5(
                                categoria_s5,
                                cliente_s5,
                                fabrica_final_arquivo,
                                ce_bri_final_arquivo,
                                endereco_final_arquivo,
                                tipo_s5,
                                ip_extraido,
                                data_s5,
                                arquivo_s5.name,
                                df_s5
                            )

                            st.success(f"{total} itens salvos no Sistema 5 para o processo {ip_extraido} ✅")

                except Exception as e:
                    st.error(f"Erro ao ler Excel do Sistema 5: {e}")

    st.divider()
    st.subheader("Estrutura salva no Sistema 5")

    processos_s5 = listar_processos_sistema5()

    if processos_s5.empty:
        st.info("Nenhum processo salvo ainda.")
    else:
        categorias = processos_s5["categoria"].dropna().unique().tolist()

        for categoria in categorias:
            with st.expander(f"📁 {categoria}", expanded=False):
                bloco_categoria = processos_s5[processos_s5["categoria"] == categoria]
                clientes = bloco_categoria["cliente_base"].dropna().unique().tolist()

                for cliente in clientes:
                    st.markdown(f"### 📁 {cliente}")
                    bloco_cliente = bloco_categoria[bloco_categoria["cliente_base"] == cliente]
                    fabricas = bloco_cliente["fabrica"].dropna().unique().tolist()

                    for fabrica in fabricas:
                        with st.expander(f"🏭 {fabrica}", expanded=False):
                            bloco_fabrica = bloco_cliente[bloco_cliente["fabrica"] == fabrica]

                            st.caption("Processos salvos nesta fábrica")

                            for _, proc in bloco_fabrica.iterrows():
                                arquivo_id = int(proc["arquivo_id"])
                                titulo_processo = (
                                    f"📄 {proc.get('tipo_processo', '')} | "
                                    f"{proc.get('ip_processo', '')} | "
                                    f"{proc.get('data_processo', '')} | "
                                    f"{proc.get('arquivo_nome', '')} | "
                                    f"{int(proc.get('qtd_itens', 0) or 0)} itens"
                                )

                                with st.expander(titulo_processo, expanded=False):
                                    colp1, colp2, colp3 = st.columns(3)

                                    colp1.metric("IP", proc.get("ip_processo", ""))
                                    colp2.metric("Tipo", proc.get("tipo_processo", ""))
                                    colp3.metric("Itens", int(proc.get("qtd_itens", 0) or 0))

                                    st.caption(f"CE-BRI: {proc.get('ce_bri', '')}")
                                    st.caption(f"Endereço: {proc.get('endereco_fabrica', '')}")

                                    termo_item_processo = st.text_input(
                                        "Pesquisar item dentro deste processo",
                                        placeholder="Digite modelo, marca, descrição ou código",
                                        key=f"pesquisar_item_processo_{arquivo_id}"
                                    )

                                    itens_do_processo = buscar_itens_processo_sistema5(
                                        arquivo_id,
                                        termo_item_processo
                                    )

                                    if itens_do_processo.empty:
                                        st.warning("Nenhum item encontrado dentro deste processo.")
                                    else:
                                        st.success(f"{len(itens_do_processo)} item(ns) encontrado(s) neste processo ✅")
                                        st.dataframe(itens_do_processo, use_container_width=True)

                                        st.download_button(
                                            "Baixar itens deste processo CSV",
                                            itens_do_processo.to_csv(index=False, sep=";").encode(
                                                "ISO-8859-1",
                                                errors="replace"
                                            ),
                                            f"sistema5_processo_{arquivo_id}.csv",
                                            "text/csv",
                                            key=f"download_itens_processo_{arquivo_id}"
                                        )

                                    st.divider()
                                    st.caption("🗑️ Remover este processo")
                                    confirmar_del_proc = st.checkbox(
                                        f"Confirmo que quero remover permanentemente este processo e seus {int(proc.get('qtd_itens', 0) or 0)} itens",
                                        key=f"confirmar_del_proc_{arquivo_id}"
                                    )
                                    if confirmar_del_proc:
                                        if st.button(
                                            "🗑️ Deletar processo",
                                            key=f"btn_del_proc_{arquivo_id}"
                                        ):
                                            try:
                                                qtd_del = deletar_processo_sistema5(arquivo_id)
                                                st.success(f"Processo {proc.get('ip_processo', '')} removido. {qtd_del} itens deletados ✅")
                                                st.rerun()
                                            except Exception as e:
                                                st.error(f"Erro ao deletar: {e}")

        st.download_button(
            "Baixar resumo Sistema 5 CSV",
            processos_s5.to_csv(index=False, sep=";").encode("ISO-8859-1", errors="replace"),
            "sistema5_resumo.csv",
            "text/csv",
            key="download_sistema5_resumo"
        )



# ==========================================
# ABA 7 - IP-BRI PENDENTES
# ==========================================

with aba7:
    st.title("IP-BRI Pendentes")

    st.info(
        "Use esta aba para cadastrar manualmente o IP-BRI de famílias que aparecem no Sistema 5, "
        "mas ainda não possuem certificado oficial emitido/cadastrado."
    )

    st.subheader("Famílias pendentes encontradas no Sistema 5")

    pendentes_ip_bri = listar_ip_bri_pendentes_sistema5()

    if pendentes_ip_bri.empty:
        st.success("Nenhuma família pendente de IP-BRI encontrada ✅")
    else:
        st.warning(f"{len(pendentes_ip_bri)} família(s) pendente(s) de IP-BRI.")
        st.dataframe(pendentes_ip_bri, use_container_width=True)

        st.download_button(
            "Baixar pendentes CSV",
            pendentes_ip_bri.to_csv(index=False, sep=";").encode("ISO-8859-1", errors="replace"),
            "ip_bri_pendentes.csv",
            "text/csv",
            key="download_ip_bri_pendentes"
        )

        st.divider()
        st.subheader("Cadastrar IP-BRI de uma família pendente")

        opcoes_pendentes = []
        for _, linha in pendentes_ip_bri.iterrows():
            label = (
                f"{linha.get('cliente_base', '')} | "
                f"{linha.get('fabrica', '')} | "
                f"{linha.get('ce_bri', '')} | "
                f"FAM {linha.get('familia', '')} | "
                f"{linha.get('ip_processo', '')}"
            )
            opcoes_pendentes.append(label)

        escolha_pendente = st.selectbox(
            "Selecione a família pendente",
            opcoes_pendentes,
            key="select_ip_bri_pendente"
        )

        idx_escolhido = opcoes_pendentes.index(escolha_pendente)
        linha_escolhida = pendentes_ip_bri.iloc[idx_escolhido]

        col_ip1, col_ip2 = st.columns(2)

        with col_ip1:
            st.text_input("CE-BRI", value=clean(linha_escolhida.get("ce_bri", "")), disabled=True, key="ip_pendente_ce_bri")
            st.text_input("Família", value=clean(linha_escolhida.get("familia", "")), disabled=True, key="ip_pendente_familia")

            registro_copiavel = clean(linha_escolhida.get("registro", ""))

            if registro_copiavel:
                st.caption("Registro para copiar:")
                st.code(registro_copiavel)

        with col_ip2:
            novo_ip_bri_manual = st.text_input("IP-BRI a cadastrar", placeholder="Ex: IP-BRI-1550-26 ou 1550-26", key="novo_ip_bri_manual")
            obs_ip_bri_manual = st.text_input("Observação", placeholder="Opcional", key="obs_ip_bri_manual")

        if st.button("Salvar IP-BRI desta família", key="salvar_ip_bri_pendente"):
            ok, mensagem = salvar_ip_bri_familia(
                linha_escolhida.get("ce_bri", ""),
                linha_escolhida.get("familia", ""),
                novo_ip_bri_manual,
                obs_ip_bri_manual
            )

            if ok:
                st.success(mensagem)
            else:
                st.error(mensagem)

    st.divider()
    st.subheader("Cadastro manual direto")

    col_manual1, col_manual2, col_manual3 = st.columns(3)

    with col_manual1:
        ce_manual = st.text_input("CE-BRI", placeholder="CE-BRI-INNAC-02484-01A", key="manual_ce_bri_ip")

    with col_manual2:
        fam_manual = st.text_input("Família", placeholder="Ex: 12", key="manual_familia_ip")

    with col_manual3:
        ip_manual = st.text_input("IP-BRI", placeholder="Ex: IP-BRI-1550-26", key="manual_ip_bri")

    obs_manual = st.text_input("Observação manual", placeholder="Opcional", key="manual_obs_ip_bri")

    if st.button("Salvar cadastro manual", key="salvar_ip_bri_manual_direto"):
        ok, mensagem = salvar_ip_bri_familia(ce_manual, fam_manual, ip_manual, obs_manual)

        if ok:
            st.success(mensagem)
        else:
            st.error(mensagem)

    st.divider()
    st.subheader("IP-BRI manuais já cadastrados")

    manuais = listar_ip_bri_manuais()

    if manuais.empty:
        st.info("Nenhum IP-BRI manual cadastrado ainda.")
    else:
        st.dataframe(manuais, use_container_width=True)

        st.download_button(
            "Baixar IP-BRI manuais CSV",
            manuais.to_csv(index=False, sep=";").encode("ISO-8859-1", errors="replace"),
            "ip_bri_manuais.csv",
            "text/csv",
            key="download_ip_bri_manuais"
        )

        st.divider()
        st.subheader("Editar ou excluir IP-BRI manual")

        opcoes_manuais = []

        for _, linha in manuais.iterrows():
            label = (
                f"ID {linha.get('id', '')} | "
                f"{linha.get('ce_bri', '')} | "
                f"FAM {linha.get('familia', '')} | "
                f"{linha.get('ip_bri', '')}"
            )
            opcoes_manuais.append(label)

        escolha_manual = st.selectbox(
            "Selecione o cadastro para editar/excluir",
            opcoes_manuais,
            key="select_editar_ip_bri_manual"
        )

        idx_manual = opcoes_manuais.index(escolha_manual)
        linha_manual = manuais.iloc[idx_manual]

        registro_manual_id = int(linha_manual.get("id"))

        col_edit1, col_edit2, col_edit3 = st.columns(3)

        with col_edit1:
            st.caption("CE-BRI vinculado")
            st.code(clean(linha_manual.get("ce_bri", "")))

        with col_edit2:
            st.caption("Família vinculada")
            st.code(clean(linha_manual.get("familia", "")))

        with col_edit3:
            novo_ip_editado = st.text_input(
                "Novo IP-BRI",
                value=clean(linha_manual.get("ip_bri", "")),
                key=f"edit_ip_bri_novo_{registro_manual_id}"
            )

        nova_obs_editada = st.text_input(
            "Observação",
            value=clean(linha_manual.get("observacao", "")),
            key=f"edit_ip_obs_nova_{registro_manual_id}"
        )

        col_btn1, col_btn2 = st.columns(2)

        with col_btn1:
            if st.button("Salvar alteração do IP-BRI", key="btn_salvar_edit_ip_manual"):
                ok, mensagem = atualizar_ip_bri_manual_por_id(
                    registro_manual_id,
                    novo_ip_editado,
                    nova_obs_editada
                )

                if ok:
                    st.success(mensagem)
                else:
                    st.error(mensagem)

        with col_btn2:
            confirmar_exclusao_ip = st.checkbox(
                "Confirmo que quero excluir este cadastro",
                key="confirmar_excluir_ip_manual"
            )

            if st.button("Excluir cadastro selecionado", key="btn_excluir_ip_manual"):
                if not confirmar_exclusao_ip:
                    st.warning("Marque a confirmação antes de excluir.")
                else:
                    ok, mensagem = excluir_ip_bri_manual_por_id(registro_manual_id)

                    if ok:
                        st.success(mensagem)
                    else:
                        st.error(mensagem)



# ==========================================
# ABA 8 - PAINEL DE COBERTURA
# ==========================================

with aba8:
    st.title("Painel de Cobertura")

    st.info(
        "Visão geral por cliente, fábrica, CE-BRI, família e quantidade de itens. "
        "Use para saber quantas fábricas existem, quantas famílias cada fábrica possui "
        "e quantos itens existem em cada família."
    )

    resumo = cobertura_resumo_geral()

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Clientes", resumo.get("clientes_registros", 0))
    c2.metric("Fábricas", resumo.get("fabricas_registros", 0))
    c3.metric("Famílias Registros", resumo.get("familias_registros", 0))
    c4.metric("Famílias Sistema 5", resumo.get("familias_sistema5", 0))
    c5.metric("Pendentes IP-BRI", resumo.get("pendentes_ip_bri", 0))

    st.divider()

    st.subheader("Resumo por fábrica")

    df_fabricas = cobertura_fabricas()

    if df_fabricas.empty:
        st.info("Nenhuma fábrica encontrada.")
    else:
        st.dataframe(df_fabricas, use_container_width=True)

        st.download_button(
            "Baixar resumo por fábrica CSV",
            df_fabricas.to_csv(index=False, sep=";").encode("ISO-8859-1", errors="replace"),
            "painel_cobertura_fabricas.csv",
            "text/csv",
            key="download_cobertura_fabricas"
        )

    st.divider()

    st.subheader("Detalhamento por família")

    if df_fabricas.empty:
        st.info("Cadastre registros ou processos no Sistema 5 para visualizar o detalhamento.")
    else:
        clientes_opcoes = ["TODOS"] + sorted(df_fabricas["cliente_base"].dropna().astype(str).unique().tolist())
        cliente_filtro = st.selectbox("Cliente", clientes_opcoes, key="cobertura_cliente")

        df_base_filtro = df_fabricas.copy()
        if cliente_filtro != "TODOS":
            df_base_filtro = df_base_filtro[df_base_filtro["cliente_base"].astype(str) == cliente_filtro]

        fabricas_opcoes = ["TODAS"] + sorted(df_base_filtro["fabrica"].dropna().astype(str).unique().tolist())
        fabrica_filtro = st.selectbox("Fábrica", fabricas_opcoes, key="cobertura_fabrica")

        ce_opcoes_base = df_base_filtro.copy()
        if fabrica_filtro != "TODAS":
            ce_opcoes_base = ce_opcoes_base[ce_opcoes_base["fabrica"].astype(str) == fabrica_filtro]

        ce_opcoes = ["TODOS"] + sorted(ce_opcoes_base["ce_bri"].dropna().astype(str).unique().tolist())
        ce_filtro = st.selectbox("CE-BRI", ce_opcoes, key="cobertura_ce_bri")

        df_familias = cobertura_familias_detalhada(
            "" if cliente_filtro == "TODOS" else cliente_filtro,
            "" if fabrica_filtro == "TODAS" else fabrica_filtro,
            "" if ce_filtro == "TODOS" else ce_filtro
        )

        if df_familias.empty:
            st.warning("Nenhuma família encontrada para os filtros selecionados.")
        else:
            st.dataframe(df_familias, use_container_width=True)

            st.download_button(
                "Baixar detalhamento por família CSV",
                df_familias.to_csv(index=False, sep=";").encode("ISO-8859-1", errors="replace"),
                "painel_cobertura_familias.csv",
                "text/csv",
                key="download_cobertura_familias"
            )

            st.divider()
            st.subheader("Itens de uma família específica")

            opcoes_familia = []
            for _, linha in df_familias.iterrows():
                label = (
                    f"{linha.get('cliente_base', '')} | "
                    f"{linha.get('fabrica', '')} | "
                    f"{linha.get('ce_bri', '')} | "
                    f"FAM {linha.get('familia', '')} | "
                    f"{linha.get('qtd_itens_sistema5', 0)} itens"
                )
                opcoes_familia.append(label)

            escolha_fam = st.selectbox(
                "Selecione uma família para ver os itens",
                opcoes_familia,
                key="select_cobertura_familia_itens"
            )

            idx_fam = opcoes_familia.index(escolha_fam)
            linha_fam = df_familias.iloc[idx_fam]

            df_itens_fam = cobertura_itens_por_familia(
                linha_fam.get("cliente_base", ""),
                linha_fam.get("fabrica", ""),
                linha_fam.get("ce_bri", ""),
                linha_fam.get("familia", "")
            )

            if df_itens_fam.empty:
                st.info("Essa família não possui itens no Sistema 5.")
            else:
                st.success(f"{len(df_itens_fam)} item(ns) encontrado(s) nesta família.")
                st.dataframe(df_itens_fam, use_container_width=True)

                st.download_button(
                    "Baixar itens da família CSV",
                    df_itens_fam.to_csv(index=False, sep=";").encode("ISO-8859-1", errors="replace"),
                    "painel_cobertura_itens_familia.csv",
                    "text/csv",
                    key="download_cobertura_itens_familia"
                )

